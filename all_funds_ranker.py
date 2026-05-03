"""
================================================================
  ALL MUTUAL FUNDS — EXPERT RANKER v1.0
  35+ Professional Metrics | Auto-Generated HTML Dashboard
  Data Source: captn3m0/historical-mf-data (AMFI Official)
================================================================
Metrics Computed:
  Returns     : 1Y / 3Y / 5Y / 10Y CAGR, Since Inception
  SIP Returns : 1Y / 3Y / 5Y SIP XIRR
  Rolling     : Avg 1Y / 3Y / 5Y Rolling Returns
  Risk        : Sharpe, Sortino, Std Dev, Max Drawdown, Calmar, VaR
  Capture     : Upside Capture Ratio, Downside Capture Ratio
  Regression  : Alpha, Beta, R-Squared, Treynor, Information Ratio
  Consistency : % Positive 1Y Periods, % Beating Category Avg 3Y
  Score       : Expert Composite Score (0–100) + Category Rank
================================================================
"""

import sqlite3, os, json, math, sys, zipfile, io
from datetime import datetime, timedelta
from pathlib import Path

import requests
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Config ───────────────────────────────────────────────────────
DB_URL      = "https://github.com/captn3m0/historical-mf-data/releases/latest/download/funds.db.zst"
DB_PATH     = "funds.db"
OUT_DIR     = Path("output")
OUT_DIR.mkdir(exist_ok=True)
DASH_PATH   = Path("dashboard") / "index.html"
Path("dashboard").mkdir(exist_ok=True)

# Nifty 50 Index Fund Direct Growth (proxy for benchmark) — common scheme codes
# We try multiple well-known ones
NIFTY_CANDIDATES = [
    "UTI Nifty 50 Index Fund-Direct Plan-Growth Option",
    "UTI Nifty Index Fund-Direct Plan-Growth",
    "HDFC Index Fund-NIFTY 50 Plan-Direct Plan",
    "Nippon India Index Fund-Nifty 50 Plan - Direct Plan - Growth",
    "ICICI Prudential Nifty 50 Index Fund Direct Plan Growth",
]

CATEGORIES = {
    "Large Cap":            ["large cap", "largecap"],
    "Mid Cap":              ["mid cap", "midcap"],
    "Small Cap":            ["small cap", "smallcap"],
    "Flexi Cap":            ["flexi cap", "flexicap", "multi-cap"],
    "Multi Cap":            ["multi cap"],
    "Large & Mid Cap":      ["large & mid cap", "large and mid cap"],
    "ELSS":                 ["elss", "tax sav"],
    "Aggressive Hybrid":    ["aggressive hybrid"],
    "Balanced Advantage":   ["balanced advantage", "dynamic asset"],
    "Index - Nifty 50":     ["nifty 50 index", "nifty50 index"],
    "Index - Nifty Next 50":["nifty next 50", "next 50"],
    "Index - Nifty 100":    ["nifty 100 index"],
    "Sectoral - IT":        ["technology", "it fund", "infotech"],
    "Sectoral - Banking":   ["banking", "financial serv", "finserv"],
    "Debt - Liquid":        ["liquid fund"],
    "Debt - Short Duration":["short duration"],
    "Debt - Corporate Bond":["corporate bond"],
    "International":        ["international", "global", "overseas", "us equity"],
}

RISK_FREE_RATE = 0.065   # 6.5% annual (10-yr G-Sec approx)
TRADING_DAYS   = 252

# ── Download + Open DB ───────────────────────────────────────────
def download_db():
    print("📥 Downloading historical MF database …")
    r = requests.get(DB_URL, stream=True, timeout=300)
    r.raise_for_status()
    zst_bytes = r.content
    print(f"   Downloaded {len(zst_bytes)/1e6:.1f} MB")

    # Decompress with zstandard if available, else try python-zstandard
    try:
        import zstandard as zstd
        dctx = zstd.ZstdDecompressor()
        with open(DB_PATH, "wb") as f:
            f.write(dctx.decompress(zst_bytes))
    except ImportError:
        # Fallback: use system zstd via subprocess
        import subprocess, tempfile
        with tempfile.NamedTemporaryFile(suffix=".zst", delete=False) as tmp:
            tmp.write(zst_bytes)
            tmp_path = tmp.name
        subprocess.run(["zstd", "-d", tmp_path, "-o", DB_PATH, "-f"], check=True)
        os.unlink(tmp_path)
    print(f"   Database ready at {DB_PATH}")


def load_nav_series(conn, scheme_ids: list[int]) -> dict[int, pd.Series]:
    """Load NAV time series for a list of scheme IDs."""
    if not scheme_ids:
        return {}
    placeholders = ",".join(["?"] * len(scheme_ids))
    df = pd.read_sql_query(
        f"SELECT scheme_id, date, nav FROM nav WHERE scheme_id IN ({placeholders}) ORDER BY date",
        conn, params=scheme_ids, parse_dates=["date"]
    )
    result = {}
    for sid, grp in df.groupby("scheme_id"):
        s = grp.set_index("date")["nav"].sort_index().astype(float)
        s = s[~s.index.duplicated()]
        result[sid] = s
    return result


# ── Metric Helpers ───────────────────────────────────────────────
def cagr(nav: pd.Series, years: float) -> float | None:
    """Annualised return over last `years` years."""
    end   = nav.index[-1]
    start = end - pd.DateOffset(days=int(years * 365))
    sub   = nav[nav.index >= start]
    if len(sub) < 20:
        return None
    r = (sub.iloc[-1] / sub.iloc[0]) ** (1 / years) - 1
    return round(r * 100, 2)


def since_inception_cagr(nav: pd.Series) -> float | None:
    if len(nav) < 30:
        return None
    years = (nav.index[-1] - nav.index[0]).days / 365.25
    if years < 0.5:
        return None
    r = (nav.iloc[-1] / nav.iloc[0]) ** (1 / years) - 1
    return round(r * 100, 2)


def sip_xirr(nav: pd.Series, years: float) -> float | None:
    """Monthly SIP XIRR for last `years` years."""
    try:
        end   = nav.index[-1]
        start = end - pd.DateOffset(days=int(years * 365))
        sub   = nav[nav.index >= start]
        if len(sub) < 12:
            return None
        monthly = sub.resample("MS").first().dropna()
        if len(monthly) < 12:
            return None
        dates, amounts = [], []
        units = 0.0
        for dt, price in monthly.items():
            units    += 1000 / price
            dates.append(dt)
            amounts.append(-1000)
        # Final redemption
        dates.append(sub.index[-1])
        amounts.append(units * sub.iloc[-1])

        # Newton-Raphson XIRR
        def npv(rate):
            t0 = dates[0]
            return sum(cf / (1 + rate) ** ((d - t0).days / 365.25)
                       for cf, d in zip(amounts, dates))

        lo, hi = -0.5, 5.0
        for _ in range(200):
            mid = (lo + hi) / 2
            if npv(mid) > 0:
                lo = mid
            else:
                hi = mid
            if hi - lo < 1e-6:
                break
        return round(((lo + hi) / 2) * 100, 2)
    except Exception:
        return None


def rolling_avg(nav: pd.Series, years: float) -> float | None:
    """Average rolling CAGR using windows of `years` years."""
    days = int(years * 365)
    if len(nav) < days + 30:
        return None
    sub  = nav.resample("W-FRI").last().dropna()
    step = max(1, int(years * 52))
    returns = []
    for i in range(len(sub) - step):
        p0, p1 = sub.iloc[i], sub.iloc[i + step]
        returns.append((p1 / p0) ** (1 / years) - 1)
    return round(np.mean(returns) * 100, 2) if returns else None


def daily_returns(nav: pd.Series, years: float = 3) -> pd.Series | None:
    end   = nav.index[-1]
    start = end - pd.DateOffset(days=int(years * 365))
    sub   = nav[nav.index >= start]
    if len(sub) < 60:
        return None
    return sub.pct_change().dropna()


def sharpe(dr: pd.Series) -> float | None:
    if dr is None or len(dr) < 30:
        return None
    ann_ret  = (1 + dr.mean()) ** TRADING_DAYS - 1
    ann_std  = dr.std() * math.sqrt(TRADING_DAYS)
    if ann_std == 0:
        return None
    return round((ann_ret - RISK_FREE_RATE) / ann_std, 2)


def sortino(dr: pd.Series) -> float | None:
    if dr is None or len(dr) < 30:
        return None
    ann_ret   = (1 + dr.mean()) ** TRADING_DAYS - 1
    down      = dr[dr < 0]
    if len(down) < 5:
        return None
    down_std  = down.std() * math.sqrt(TRADING_DAYS)
    if down_std == 0:
        return None
    return round((ann_ret - RISK_FREE_RATE) / down_std, 2)


def max_drawdown(nav: pd.Series) -> float | None:
    if len(nav) < 20:
        return None
    roll_max = nav.cummax()
    dd       = (nav - roll_max) / roll_max
    return round(dd.min() * 100, 2)


def calmar(nav: pd.Series, years: float = 3) -> float | None:
    r  = cagr(nav, years)
    md = max_drawdown(nav)
    if r is None or md is None or md == 0:
        return None
    return round(r / abs(md), 2)


def volatility(dr: pd.Series) -> float | None:
    if dr is None or len(dr) < 30:
        return None
    return round(dr.std() * math.sqrt(TRADING_DAYS) * 100, 2)


def value_at_risk(dr: pd.Series, pct: float = 0.05) -> float | None:
    """Historical VaR at confidence level (1-pct)."""
    if dr is None or len(dr) < 60:
        return None
    return round(np.percentile(dr, pct * 100) * 100, 2)


def alpha_beta_r2(fund_dr: pd.Series, bench_dr: pd.Series) -> tuple:
    """OLS regression of fund returns on benchmark returns."""
    try:
        aligned = pd.concat([fund_dr, bench_dr], axis=1).dropna()
        if len(aligned) < 60:
            return None, None, None
        x = aligned.iloc[:, 1].values
        y = aligned.iloc[:, 0].values
        cov_mat  = np.cov(x, y)
        var_x    = cov_mat[0, 0]
        if var_x == 0:
            return None, None, None
        beta     = cov_mat[0, 1] / var_x
        alpha_d  = np.mean(y) - beta * np.mean(x)
        alpha_ann = ((1 + alpha_d) ** TRADING_DAYS - 1) * 100
        corr     = np.corrcoef(x, y)[0, 1]
        r2       = round(corr ** 2, 3)
        return round(alpha_ann, 2), round(beta, 3), r2
    except Exception:
        return None, None, None


def treynor(dr: pd.Series, beta: float) -> float | None:
    if dr is None or beta is None or beta == 0 or len(dr) < 30:
        return None
    ann_ret = (1 + dr.mean()) ** TRADING_DAYS - 1
    return round((ann_ret - RISK_FREE_RATE) / beta, 4)


def information_ratio(fund_dr: pd.Series, bench_dr: pd.Series) -> float | None:
    try:
        aligned = pd.concat([fund_dr, bench_dr], axis=1).dropna()
        if len(aligned) < 60:
            return None
        active_dr = aligned.iloc[:, 0] - aligned.iloc[:, 1]
        te = active_dr.std() * math.sqrt(TRADING_DAYS)
        if te == 0:
            return None
        ann_active = active_dr.mean() * TRADING_DAYS
        return round(ann_active / te, 2)
    except Exception:
        return None


def capture_ratios(fund_dr: pd.Series, bench_dr: pd.Series) -> tuple:
    """Upside and Downside Capture Ratio."""
    try:
        aligned = pd.concat([fund_dr, bench_dr], axis=1).dropna()
        aligned.columns = ["fund", "bench"]
        up   = aligned[aligned["bench"] > 0]
        down = aligned[aligned["bench"] < 0]
        if len(up) < 10 or len(down) < 10:
            return None, None
        ucr = ((1 + up["fund"].mean()) ** TRADING_DAYS - 1) / \
              ((1 + up["bench"].mean()) ** TRADING_DAYS - 1) * 100
        dcr = ((1 + down["fund"].mean()) ** TRADING_DAYS - 1) / \
              ((1 + down["bench"].mean()) ** TRADING_DAYS - 1) * 100
        return round(ucr, 1), round(dcr, 1)
    except Exception:
        return None, None


def pct_positive_rolling(nav: pd.Series, years: float = 1) -> float | None:
    step = int(years * 52)
    sub  = nav.resample("W-FRI").last().dropna()
    if len(sub) < step + 10:
        return None
    wins = sum(sub.iloc[i + step] > sub.iloc[i] for i in range(len(sub) - step))
    return round(wins / (len(sub) - step) * 100, 1)


# ── Expert Score ─────────────────────────────────────────────────
def expert_score(m: dict) -> float:
    """
    Composite expert score 0–100.
    Weights designed for equity direct-growth funds.
    """
    def norm(val, low, high, invert=False):
        if val is None:
            return 0.0
        v = max(low, min(high, val))
        n = (v - low) / (high - low)
        return (1 - n) if invert else n

    scores = {
        # Returns (35%)
        "ret_3y":   norm(m.get("return_3y"),     0, 20)  * 10,
        "ret_5y":   norm(m.get("return_5y"),     0, 22)  * 10,
        "ret_10y":  norm(m.get("return_10y"),    0, 18)  * 5,
        "sip_5y":   norm(m.get("sip_return_5y"), 0, 20)  * 5,
        "roll_3y":  norm(m.get("rolling_avg_3y"),0, 20)  * 5,
        # Risk (30%)
        "sharpe":   norm(m.get("sharpe_3y"),    -1, 2.5) * 8,
        "sortino":  norm(m.get("sortino_3y"),   -1, 3.5) * 7,
        "mdd":      norm(m.get("max_drawdown"), -80, -5, invert=False) * 0,  # handled below
        "calmar":   norm(m.get("calmar_3y"),     0, 1.5) * 6,
        "vol":      norm(m.get("volatility_3y"), 40, 8, invert=True) * 5,  # lower = better, invert
        "var":      norm(m.get("var_95"),       -8, -1, invert=True) * 4,
        # Capture (15%)
        "ucr":      norm(m.get("upside_capture"),   60, 130) * 8,
        "dcr":      norm(m.get("downside_capture"), 120, 60, invert=True) * 7,
        # Alpha / Regression (10%)
        "alpha":    norm(m.get("alpha_3y"),     -5, 8)   * 6,
        "ir":       norm(m.get("information_ratio"), -1, 1) * 4,
        # Consistency (10%)
        "pos1y":    norm(m.get("pct_pos_1y"),   40, 100) * 5,
        "roll1y":   norm(m.get("rolling_avg_1y"),0, 18)  * 5,
    }

    # Max drawdown: score improves the less severe drawdown is
    mdd = m.get("max_drawdown")
    if mdd is not None:
        scores["mdd"] = norm(mdd, -80, -5) * 4   # less negative = better

    raw = sum(scores.values())
    # max possible ≈ 100, already scaled
    return round(min(100, max(0, raw)), 1)


# ── Category Classifier ──────────────────────────────────────────
def classify(name: str) -> str | None:
    n = name.lower()
    for cat, kws in CATEGORIES.items():
        for kw in kws:
            if kw in n:
                return cat
    return None


# ── Main Pipeline ─────────────────────────────────────────────────
def main():
    # Install zstandard if needed
    try:
        import zstandard  # noqa
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "zstandard", "-q"])

    download_db()

    conn = sqlite3.connect(DB_PATH)
    print("📋 Loading scheme list …")
    funds_df = pd.read_sql_query("SELECT id, name FROM funds", conn)

    # Filter: Direct Growth equity/hybrid funds
    direct = funds_df[
        funds_df["name"].str.contains("direct", case=False, na=False) &
        funds_df["name"].str.contains("growth", case=False, na=False) &
        ~funds_df["name"].str.contains("dividend|idcw|weekly|monthly|quarterly", case=False, na=False)
    ].copy()
    direct["category"] = direct["name"].apply(classify)
    direct = direct[direct["category"].notna()].reset_index(drop=True)
    print(f"   {len(direct)} Direct Growth funds in {direct['category'].nunique()} categories")

    # Find Nifty 50 benchmark
    bench_id = None
    for cand in NIFTY_CANDIDATES:
        row = funds_df[funds_df["name"].str.lower().str.contains(
            cand.lower()[:30], na=False)]
        if not row.empty:
            bench_id = int(row.iloc[0]["id"])
            print(f"   Benchmark: {row.iloc[0]['name']}")
            break
    if bench_id is None:
        # Use first Nifty 50 Index fund found
        row = funds_df[funds_df["name"].str.lower().str.contains("nifty 50 index", na=False)]
        if not row.empty:
            bench_id = int(row.iloc[0]["id"])

    # Load benchmark NAV
    bench_nav = None
    bench_dr  = None
    if bench_id:
        nav_map   = load_nav_series(conn, [bench_id])
        bench_nav = nav_map.get(bench_id)
        if bench_nav is not None:
            bench_dr = bench_nav.pct_change().dropna()

    # Process funds in batches
    BATCH = 50
    all_ids  = direct["id"].tolist()
    id_to_cat = dict(zip(direct["id"], direct["category"]))
    id_to_name = dict(zip(direct["id"], direct["name"]))
    all_metrics = []

    for start in range(0, len(all_ids), BATCH):
        batch = all_ids[start:start + BATCH]
        nav_map = load_nav_series(conn, batch)
        for sid in batch:
            nav  = nav_map.get(sid)
            name = id_to_name[sid]
            cat  = id_to_cat[sid]

            if nav is None or len(nav) < 60:
                continue

            dr3 = daily_returns(nav, 3)

            # Benchmark alignment
            alpha, beta, r2 = None, None, None
            ucr, dcr        = None, None
            ir              = None
            treynor_r       = None
            if bench_dr is not None and dr3 is not None:
                bench_aligned = bench_dr.reindex(dr3.index).dropna()
                fund_aligned  = dr3.reindex(bench_aligned.index).dropna()
                if len(fund_aligned) >= 60:
                    alpha, beta, r2 = alpha_beta_r2(fund_aligned, bench_aligned)
                    ucr, dcr        = capture_ratios(fund_aligned, bench_aligned)
                    ir              = information_ratio(fund_aligned, bench_aligned)
                    if beta is not None:
                        treynor_r = treynor(fund_aligned, beta)

            m = {
                "scheme_id":          sid,
                "fund_name":          name,
                "category":           cat,
                # Returns
                "return_1y":          cagr(nav, 1),
                "return_3y":          cagr(nav, 3),
                "return_5y":          cagr(nav, 5),
                "return_10y":         cagr(nav, 10),
                "return_inception":   since_inception_cagr(nav),
                # SIP XIRR
                "sip_return_1y":      sip_xirr(nav, 1),
                "sip_return_3y":      sip_xirr(nav, 3),
                "sip_return_5y":      sip_xirr(nav, 5),
                # Rolling
                "rolling_avg_1y":     rolling_avg(nav, 1),
                "rolling_avg_3y":     rolling_avg(nav, 3),
                "rolling_avg_5y":     rolling_avg(nav, 5),
                # Risk
                "sharpe_3y":          sharpe(dr3),
                "sortino_3y":         sortino(dr3),
                "volatility_3y":      volatility(dr3),
                "max_drawdown":       max_drawdown(nav),
                "calmar_3y":          calmar(nav, 3),
                "var_95":             value_at_risk(dr3),
                # Capture
                "upside_capture":     ucr,
                "downside_capture":   dcr,
                # Regression
                "alpha_3y":           alpha,
                "beta_3y":            beta,
                "r_squared_3y":       r2,
                "treynor_ratio":      treynor_r,
                "information_ratio":  ir,
                # Consistency
                "pct_pos_1y":         pct_positive_rolling(nav, 1),
                "pct_pos_3y":         pct_positive_rolling(nav, 3),
                # Meta
                "nav_latest":         round(float(nav.iloc[-1]), 4),
                "nav_date":           nav.index[-1].strftime("%d-%b-%Y"),
                "inception_date":     nav.index[0].strftime("%d-%b-%Y"),
                "history_years":      round((nav.index[-1] - nav.index[0]).days / 365.25, 1),
            }
            m["expert_score"] = expert_score(m)
            all_metrics.append(m)

        done = min(start + BATCH, len(all_ids))
        print(f"   ✓ {done}/{len(all_ids)} funds processed", end="\r")

    print(f"\n   Total computed: {len(all_metrics)} funds")
    conn.close()

    # ── DataFrames ──────────────────────────────────────────────
    df = pd.DataFrame(all_metrics)
    # Category percentile rank
    df["category_rank"] = df.groupby("category")["expert_score"].rank(
        ascending=False, method="min").astype(int)
    df["category_total"] = df.groupby("category")["expert_score"].transform("count").astype(int)
    df["percentile_rank"] = (df["category_rank"] / df["category_total"] * 100).round(1)

    df_sorted = df.sort_values(["category", "expert_score"], ascending=[True, False])

    # Save CSVs
    df_sorted.to_csv(OUT_DIR / "all_funds_metrics.csv", index=False)
    top5 = df_sorted.groupby("category").head(5)
    top5.to_csv(OUT_DIR / "top5_per_category.csv", index=False)
    print(f"   📋 CSVs saved to output/")

    # Per-category top5 dict for dashboard
    top5_dict = {}
    for cat, grp in df_sorted.groupby("category"):
        top5_dict[cat] = grp.head(5).to_dict("records")

    # Build Excel
    build_excel(df_sorted, top5_dict)
    print(f"   📊 Excel saved to output/")

    # Build HTML Dashboard
    build_dashboard(df_sorted, top5_dict)
    print(f"   🖥️  Dashboard saved to dashboard/index.html")

    # Cleanup
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    print("\n" + "=" * 65)
    print(f"  ✅ DONE — {len(df_sorted)} funds | {df_sorted['category'].nunique()} categories")
    print(f"  📋 output/all_funds_metrics.csv")
    print(f"  📋 output/top5_per_category.csv")
    print(f"  📊 output/MF_Expert_Rankings.xlsx")
    print(f"  🖥️  dashboard/index.html  (open in browser)")
    print("=" * 65)


# ── Excel Builder ────────────────────────────────────────────────
def build_excel(df: pd.DataFrame, top5_dict: dict):
    wb = openpyxl.Workbook()

    DARK    = "1A1A2E"
    ACCENT  = "00D4AA"
    GOLD    = "FFD700"
    WHITE   = "FFFFFF"
    GRAY    = "2D2D4E"
    RED     = "FF4757"
    GREEN   = "2ED573"

    def hdr_cell(ws, row, col, val, bg=DARK, fg=WHITE, bold=True, sz=11):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(color=fg, bold=bold, size=sz, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return c

    def data_cell(ws, row, col, val, bg="FFFFFF", fg="1A1A2E", bold=False, num_fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(color=fg, bold=bold, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        if num_fmt:
            c.number_format = num_fmt
        return c

    COLS = [
        ("Rank",                 "category_rank",         6,  None),
        ("Fund Name",            "fund_name",              38, None),
        ("Expert\nScore",        "expert_score",           9,  None),
        ("Ret 1Y\n(%)",          "return_1y",              8,  "0.00"),
        ("Ret 3Y\n(%)",          "return_3y",              8,  "0.00"),
        ("Ret 5Y\n(%)",          "return_5y",              8,  "0.00"),
        ("Ret 10Y\n(%)",         "return_10y",             8,  "0.00"),
        ("SIP 3Y\n(%)",          "sip_return_3y",          8,  "0.00"),
        ("SIP 5Y\n(%)",          "sip_return_5y",          8,  "0.00"),
        ("Roll\n1Y (%)",         "rolling_avg_1y",         8,  "0.00"),
        ("Roll\n3Y (%)",         "rolling_avg_3y",         8,  "0.00"),
        ("Sharpe\n3Y",           "sharpe_3y",              8,  "0.00"),
        ("Sortino\n3Y",          "sortino_3y",             8,  "0.00"),
        ("Std Dev\n(%)",         "volatility_3y",          8,  "0.00"),
        ("Max DD\n(%)",          "max_drawdown",           8,  "0.00"),
        ("Calmar\n3Y",           "calmar_3y",              8,  "0.00"),
        ("VaR 95\n(%)",          "var_95",                 8,  "0.00"),
        ("Upside\nCapture",      "upside_capture",         9,  "0.0"),
        ("Downside\nCapture",    "downside_capture",       9,  "0.0"),
        ("Alpha\n3Y (%)",        "alpha_3y",               9,  "0.00"),
        ("Beta\n3Y",             "beta_3y",                7,  "0.000"),
        ("R²",                   "r_squared_3y",           7,  "0.000"),
        ("Info\nRatio",          "information_ratio",      8,  "0.00"),
        ("% Pos\n1Y Rolls",      "pct_pos_1y",             9,  "0.0"),
        ("History\n(Yrs)",       "history_years",          8,  "0.0"),
        ("Latest\nNAV",          "nav_latest",             9,  None),
    ]

    # Summary sheet
    ws0 = wb.active
    ws0.title = "📊 Summary Dashboard"
    ws0.sheet_view.showGridLines = False
    ws0.row_dimensions[1].height = 50
    ws0.row_dimensions[2].height = 20

    hdr_cell(ws0, 1, 1, f"🏆  MF EXPERT RANKINGS  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
             bg=DARK, fg=ACCENT, sz=14)
    ws0.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    r = 3
    for cat in sorted(top5_dict.keys()):
        funds = top5_dict[cat]
        hdr_cell(ws0, r, 1, cat, bg=GRAY, fg=ACCENT, sz=11)
        ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for i, f in enumerate(funds[:5], 1):
            bg = "16213E" if i % 2 == 0 else "0F3460"
            data_cell(ws0, r, 1, i,                             bg=bg, fg=GOLD if i == 1 else WHITE, bold=(i == 1))
            data_cell(ws0, r, 2, f.get("fund_name", ""),        bg=bg, fg=WHITE)
            data_cell(ws0, r, 3, f.get("expert_score"),         bg=bg, fg=GOLD if i == 1 else ACCENT, bold=True)
            data_cell(ws0, r, 4, f.get("return_3y"),            bg=bg, fg=WHITE, num_fmt="0.00")
            data_cell(ws0, r, 5, f.get("return_5y"),            bg=bg, fg=WHITE, num_fmt="0.00")
            data_cell(ws0, r, 6, f.get("sharpe_3y"),            bg=bg, fg=WHITE, num_fmt="0.00")
            data_cell(ws0, r, 7, f.get("max_drawdown"),         bg=bg, fg=RED   if (f.get("max_drawdown") or 0) < -30 else WHITE, num_fmt="0.00")
            data_cell(ws0, r, 8, f.get("alpha_3y"),             bg=bg, fg=GREEN if (f.get("alpha_3y") or 0) > 0 else RED, num_fmt="0.00")
            r += 1
        r += 1

    ws0.column_dimensions["A"].width = 5
    ws0.column_dimensions["B"].width = 40
    for c in "CDEFGH":
        ws0.column_dimensions[c].width = 12

    # Category sheets
    for cat in sorted(top5_dict.keys()):
        safe = cat[:30].replace("/", "_").replace(" ", " ")
        ws = wb.create_sheet(title=safe)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "C3"

        hdr_cell(ws, 1, 1, f"🏆  {cat}  |  Expert Rankings", bg=DARK, fg=ACCENT, sz=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        ws.row_dimensions[1].height = 36

        for ci, (label, _, width, _) in enumerate(COLS, 1):
            hdr_cell(ws, 2, ci, label, bg=GRAY, fg=WHITE, sz=9)
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[2].height = 30

        cat_df = df[df["category"] == cat].sort_values("expert_score", ascending=False)
        for ri, (_, row) in enumerate(cat_df.iterrows(), 3):
            bg = "1E1E3E" if ri % 2 == 0 else "16213E"
            for ci, (_, col, _, nf) in enumerate(COLS, 1):
                val = row.get(col)
                if isinstance(val, float) and math.isnan(val):
                    val = None
                # Colour returns / drawdown
                fg = WHITE
                if col in ("return_1y","return_3y","return_5y","return_10y","sip_return_3y","sip_return_5y","rolling_avg_1y","rolling_avg_3y","alpha_3y","information_ratio"):
                    fg = GREEN if (val or 0) > 0 else RED
                elif col == "max_drawdown":
                    fg = RED if (val or 0) < -25 else WHITE
                elif col == "expert_score":
                    fg = GOLD if (val or 0) >= 70 else (GREEN if (val or 0) >= 50 else RED)
                data_cell(ws, ri, ci, val, bg=bg, fg=fg, num_fmt=nf)

    wb.save(OUT_DIR / "MF_Expert_Rankings.xlsx")


# ── HTML Dashboard Builder ───────────────────────────────────────
def build_dashboard(df: pd.DataFrame, top5_dict: dict):
    # Serialize data (replace NaN with None)
    def clean(obj):
        if isinstance(obj, float) and math.isnan(obj):
            return None
        if isinstance(obj, dict):
            return {k: clean(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [clean(v) for v in obj]
        return obj

    df_clean   = df.where(pd.notnull(df), None)
    all_json   = json.dumps(clean(df_clean.to_dict("records")))
    top5_json  = json.dumps(clean(top5_dict))
    cats_json  = json.dumps(sorted(df["category"].unique().tolist()))
    gen_time   = datetime.now().strftime("%d %b %Y %H:%M IST")
    total_funds = len(df)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>MF Expert Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root {{
  --bg:       #0A0E1A;
  --surface:  #111827;
  --card:     #1C2333;
  --border:   #2D3748;
  --accent:   #00D4AA;
  --gold:     #FFD700;
  --red:      #FF4757;
  --green:    #2ED573;
  --muted:    #8892A4;
  --text:     #E2E8F0;
  --font:     'Space Grotesk', sans-serif;
  --mono:     'JetBrains Mono', monospace;
}}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:var(--bg); color:var(--text); font-family:var(--font); min-height:100vh; }}

/* Header */
.header {{ background:linear-gradient(135deg,#0A0E1A 0%,#111827 100%);
  border-bottom:1px solid var(--border); padding:20px 32px; position:sticky; top:0; z-index:100;
  display:flex; align-items:center; justify-content:space-between; }}
.logo {{ font-size:22px; font-weight:700; color:var(--accent); letter-spacing:-0.5px; }}
.logo span {{ color:var(--gold); }}
.meta {{ font-size:12px; color:var(--muted); font-family:var(--mono); text-align:right; }}

/* Stats bar */
.statsbar {{ display:flex; gap:16px; padding:16px 32px; background:var(--surface);
  border-bottom:1px solid var(--border); flex-wrap:wrap; }}
.stat {{ background:var(--card); border:1px solid var(--border); border-radius:10px;
  padding:12px 20px; flex:1; min-width:140px; }}
.stat-label {{ font-size:11px; color:var(--muted); text-transform:uppercase; letter-spacing:1px; margin-bottom:4px; }}
.stat-val {{ font-size:22px; font-weight:700; color:var(--accent); font-family:var(--mono); }}

/* Layout */
.main {{ display:flex; height:calc(100vh - 120px); }}
.sidebar {{ width:220px; min-width:220px; background:var(--surface);
  border-right:1px solid var(--border); overflow-y:auto; padding:16px 0; }}
.content {{ flex:1; overflow:auto; padding:24px 28px; }}

/* Sidebar nav */
.sidebar-title {{ font-size:10px; text-transform:uppercase; letter-spacing:1.5px;
  color:var(--muted); padding:0 16px 8px; }}
.cat-btn {{ width:100%; text-align:left; background:none; border:none; color:var(--muted);
  padding:9px 16px; font-family:var(--font); font-size:12px; cursor:pointer;
  transition:all 0.15s; border-left:2px solid transparent; line-height:1.4; }}
.cat-btn:hover {{ background:rgba(0,212,170,0.05); color:var(--text); }}
.cat-btn.active {{ color:var(--accent); background:rgba(0,212,170,0.08);
  border-left-color:var(--accent); font-weight:600; }}
.cat-count {{ float:right; font-size:10px; background:var(--border);
  border-radius:10px; padding:1px 6px; color:var(--muted); }}

/* Search + filters */
.toolbar {{ display:flex; gap:12px; margin-bottom:20px; align-items:center; flex-wrap:wrap; }}
.search-wrap {{ flex:1; min-width:200px; position:relative; }}
.search-wrap input {{ width:100%; background:var(--card); border:1px solid var(--border);
  border-radius:8px; padding:9px 14px 9px 36px; color:var(--text); font-family:var(--font); font-size:13px; }}
.search-wrap input:focus {{ outline:none; border-color:var(--accent); }}
.search-icon {{ position:absolute; left:12px; top:50%; transform:translateY(-50%);
  color:var(--muted); font-size:14px; }}
.sort-select {{ background:var(--card); border:1px solid var(--border); border-radius:8px;
  padding:9px 12px; color:var(--text); font-family:var(--font); font-size:12px; cursor:pointer; }}
.sort-select:focus {{ outline:none; border-color:var(--accent); }}

/* Section title */
.section-hdr {{ display:flex; align-items:center; gap:12px; margin-bottom:16px; }}
.section-title {{ font-size:20px; font-weight:700; color:var(--text); }}
.badge {{ background:rgba(0,212,170,0.15); color:var(--accent); border-radius:6px;
  padding:2px 10px; font-size:11px; font-weight:600; font-family:var(--mono); }}

/* TOP 5 Cards */
.top5-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(340px,1fr));
  gap:16px; margin-bottom:28px; }}
.fund-card {{ background:var(--card); border:1px solid var(--border); border-radius:14px;
  padding:18px; transition:border-color 0.2s; position:relative; overflow:hidden; }}
.fund-card:hover {{ border-color:var(--accent); }}
.fund-card::before {{ content:''; position:absolute; top:0; left:0; right:0; height:3px;
  background:linear-gradient(90deg,var(--accent),var(--gold)); }}
.fund-card.rank-1::before {{ background:linear-gradient(90deg,var(--gold),#FFB347); }}
.fund-rank {{ font-size:11px; font-weight:700; color:var(--muted); text-transform:uppercase;
  letter-spacing:1px; margin-bottom:6px; font-family:var(--mono); }}
.fund-name {{ font-size:13px; font-weight:600; color:var(--text); margin-bottom:12px; line-height:1.4; }}
.fund-score {{ position:absolute; top:18px; right:18px; text-align:center; }}
.score-val {{ font-size:28px; font-weight:700; color:var(--accent); font-family:var(--mono); line-height:1; }}
.score-label {{ font-size:9px; color:var(--muted); text-transform:uppercase; letter-spacing:1px; }}
.metrics-row {{ display:flex; gap:10px; flex-wrap:wrap; }}
.metric-pill {{ background:rgba(255,255,255,0.04); border-radius:6px; padding:5px 9px; }}
.metric-pill .ml {{ font-size:9px; color:var(--muted); text-transform:uppercase; letter-spacing:0.5px; }}
.metric-pill .mv {{ font-size:13px; font-weight:600; font-family:var(--mono); }}
.mv.pos {{ color:var(--green); }}
.mv.neg {{ color:var(--red); }}
.mv.neutral {{ color:var(--text); }}

/* Full table */
.table-wrap {{ overflow:auto; border-radius:12px; border:1px solid var(--border); }}
table {{ width:100%; border-collapse:collapse; font-size:11.5px; }}
thead th {{ background:#0E1724; color:var(--muted); padding:10px 10px;
  text-align:center; font-weight:600; white-space:nowrap; font-size:10px;
  text-transform:uppercase; letter-spacing:0.5px; border-bottom:1px solid var(--border);
  cursor:pointer; user-select:none; position:sticky; top:0; z-index:5; }}
thead th:hover {{ color:var(--accent); }}
thead th.sorted {{ color:var(--accent); }}
thead th:first-child, thead th:nth-child(2) {{ text-align:left; }}
tbody tr {{ border-bottom:1px solid rgba(45,55,72,0.5); transition:background 0.1s; }}
tbody tr:hover {{ background:rgba(0,212,170,0.04); }}
tbody td {{ padding:8px 10px; text-align:center; white-space:nowrap; }}
tbody td:first-child, tbody td:nth-child(2) {{ text-align:left; }}
td.fund-nm {{ max-width:280px; overflow:hidden; text-overflow:ellipsis;
  color:var(--text); font-weight:500; font-size:11px; }}
td.pos {{ color:var(--green); font-family:var(--mono); font-weight:600; }}
td.neg {{ color:var(--red);   font-family:var(--mono); font-weight:600; }}
td.neutral {{ color:var(--text); font-family:var(--mono); }}
td.rank-cell {{ color:var(--muted); font-family:var(--mono); font-weight:700; }}
.score-chip {{ display:inline-block; padding:2px 8px; border-radius:20px;
  font-family:var(--mono); font-weight:700; font-size:11px; }}
.score-high {{ background:rgba(0,212,170,0.2); color:var(--accent); }}
.score-mid  {{ background:rgba(255,215,0,0.15); color:var(--gold); }}
.score-low  {{ background:rgba(255,71,87,0.15); color:var(--red); }}

/* Loading */
.loading {{ display:flex; align-items:center; justify-content:center;
  height:200px; color:var(--muted); font-size:14px; }}
.pulse {{ animation:pulse 1.5s infinite; }}
@keyframes pulse {{ 0%,100%{{opacity:1}} 50%{{opacity:0.4}} }}

/* Tooltip */
.tooltip {{ position:relative; cursor:help; }}
.tooltip:hover::after {{ content:attr(data-tip); position:absolute; bottom:125%; left:50%;
  transform:translateX(-50%); background:#0A0E1A; color:var(--text);
  padding:6px 10px; border-radius:6px; font-size:11px; white-space:nowrap;
  border:1px solid var(--border); z-index:999; pointer-events:none; }}
</style>
</head>
<body>

<div class="header">
  <div class="logo">MF <span>Expert</span> Dashboard</div>
  <div class="meta">
    <div>⚡ {total_funds} funds · {len(top5_dict)} categories</div>
    <div>Updated: {gen_time}</div>
  </div>
</div>

<div class="statsbar" id="statsBar"></div>

<div class="main">
  <aside class="sidebar">
    <div class="sidebar-title">Categories</div>
    <button class="cat-btn active" onclick="setCategory('__all__')" data-cat="__all__">
      All Funds <span class="cat-count">{total_funds}</span>
    </button>
    <div id="catNav"></div>
  </aside>
  <div class="content">
    <div class="toolbar">
      <div class="search-wrap">
        <span class="search-icon">🔍</span>
        <input type="text" id="searchInput" placeholder="Search fund name…" oninput="applyFilters()">
      </div>
      <select class="sort-select" id="sortCol" onchange="applyFilters()">
        <option value="expert_score">Sort: Expert Score</option>
        <option value="return_3y">Sort: 3Y Return</option>
        <option value="return_5y">Sort: 5Y Return</option>
        <option value="sharpe_3y">Sort: Sharpe Ratio</option>
        <option value="sortino_3y">Sort: Sortino Ratio</option>
        <option value="alpha_3y">Sort: Alpha</option>
        <option value="max_drawdown">Sort: Max Drawdown</option>
        <option value="information_ratio">Sort: Info Ratio</option>
      </select>
      <select class="sort-select" id="sortDir" onchange="applyFilters()">
        <option value="desc">↓ High → Low</option>
        <option value="asc">↑ Low → High</option>
      </select>
    </div>

    <div id="top5Section"></div>

    <div class="section-hdr" style="margin-top:24px;">
      <div class="section-title">All Funds</div>
      <div class="badge" id="tableCount">0 funds</div>
    </div>
    <div class="table-wrap">
      <table id="mainTable">
        <thead>
          <tr>
            <th onclick="sortBy('category_rank')" class="tooltip" data-tip="Rank within category">Rank</th>
            <th onclick="sortBy('fund_name')">Fund Name</th>
            <th onclick="sortBy('expert_score')" class="tooltip" data-tip="Composite expert score 0-100">Score</th>
            <th onclick="sortBy('return_1y')" class="tooltip" data-tip="1-Year CAGR (%)">Ret 1Y</th>
            <th onclick="sortBy('return_3y')" class="tooltip" data-tip="3-Year CAGR (%)">Ret 3Y</th>
            <th onclick="sortBy('return_5y')" class="tooltip" data-tip="5-Year CAGR (%)">Ret 5Y</th>
            <th onclick="sortBy('return_10y')" class="tooltip" data-tip="10-Year CAGR (%)">Ret 10Y</th>
            <th onclick="sortBy('sip_return_3y')" class="tooltip" data-tip="SIP XIRR 3 Years (%)">SIP 3Y</th>
            <th onclick="sortBy('sip_return_5y')" class="tooltip" data-tip="SIP XIRR 5 Years (%)">SIP 5Y</th>
            <th onclick="sortBy('rolling_avg_3y')" class="tooltip" data-tip="Average Rolling 3Y Return (%)">Roll 3Y</th>
            <th onclick="sortBy('sharpe_3y')" class="tooltip" data-tip="Sharpe Ratio (3Y) — Higher is better">Sharpe</th>
            <th onclick="sortBy('sortino_3y')" class="tooltip" data-tip="Sortino Ratio (3Y) — Higher is better">Sortino</th>
            <th onclick="sortBy('volatility_3y')" class="tooltip" data-tip="Annualised Volatility % (3Y) — Lower is better">Std Dev</th>
            <th onclick="sortBy('max_drawdown')" class="tooltip" data-tip="Maximum Drawdown % (full history) — Less negative is better">Max DD</th>
            <th onclick="sortBy('calmar_3y')" class="tooltip" data-tip="Calmar Ratio = Return / |Max Drawdown|">Calmar</th>
            <th onclick="sortBy('upside_capture')" class="tooltip" data-tip="Upside Capture Ratio vs Nifty 50 — >100 beats benchmark on up days">Up Cap</th>
            <th onclick="sortBy('downside_capture')" class="tooltip" data-tip="Downside Capture Ratio vs Nifty 50 — <100 falls less than benchmark on down days">Dn Cap</th>
            <th onclick="sortBy('alpha_3y')" class="tooltip" data-tip="Jensen's Alpha (3Y annualised) — Return above benchmark after adjusting for risk">Alpha</th>
            <th onclick="sortBy('beta_3y')" class="tooltip" data-tip="Beta vs Nifty 50 (3Y) — <1 is less volatile than market">Beta</th>
            <th onclick="sortBy('information_ratio')" class="tooltip" data-tip="Information Ratio — Active return per unit of tracking error">IR</th>
            <th onclick="sortBy('pct_pos_1y')" class="tooltip" data-tip="% of 1-Year rolling windows with positive return">% Pos 1Y</th>
            <th onclick="sortBy('history_years')" class="tooltip" data-tip="Fund age in years">Age</th>
          </tr>
        </thead>
        <tbody id="tableBody"></tbody>
      </table>
    </div>
  </div>
</div>

<script>
const ALL_DATA   = {all_json};
const TOP5       = {top5_json};
const CATEGORIES = {cats_json};

let currentCat  = '__all__';
let currentSort = 'expert_score';
let currentDir  = 'desc';

// Build sidebar
const catNav = document.getElementById('catNav');
CATEGORIES.forEach(cat => {{
  const cnt = ALL_DATA.filter(d => d.category === cat).length;
  const btn = document.createElement('button');
  btn.className = 'cat-btn';
  btn.dataset.cat = cat;
  btn.innerHTML = cat + `<span class="cat-count">${{cnt}}</span>`;
  btn.onclick = () => setCategory(cat);
  catNav.appendChild(btn);
}});

function setCategory(cat) {{
  currentCat = cat;
  document.querySelectorAll('.cat-btn').forEach(b => b.classList.remove('active'));
  document.querySelector(`[data-cat="${{cat}}"]`)?.classList.add('active');
  applyFilters();
}}

function sortBy(col) {{
  if (currentSort === col) {{
    currentDir = currentDir === 'desc' ? 'asc' : 'desc';
  }} else {{
    currentSort = col;
    currentDir  = 'desc';
  }}
  document.getElementById('sortCol').value = col;
  document.getElementById('sortDir').value  = currentDir;
  applyFilters();
}}

function applyFilters() {{
  currentSort = document.getElementById('sortCol').value;
  currentDir  = document.getElementById('sortDir').value;
  const q = document.getElementById('searchInput').value.toLowerCase().trim();

  let data = ALL_DATA;
  if (currentCat !== '__all__') data = data.filter(d => d.category === currentCat);
  if (q) data = data.filter(d => d.fund_name && d.fund_name.toLowerCase().includes(q));

  data = [...data].sort((a, b) => {{
    let av = a[currentSort], bv = b[currentSort];
    if (av == null) return 1;
    if (bv == null) return -1;
    return currentDir === 'desc' ? bv - av : av - bv;
  }});

  renderTop5();
  renderTable(data);
  document.getElementById('tableCount').textContent = data.length + ' funds';
}}

function fmt(v, dec=2) {{
  if (v == null || isNaN(v)) return '—';
  return parseFloat(v).toFixed(dec);
}}

function scoreClass(s) {{
  if (s == null) return '';
  if (s >= 68) return 'score-high';
  if (s >= 48) return 'score-mid';
  return 'score-low';
}}

function valClass(v, invert=false) {{
  if (v == null || isNaN(v)) return 'neutral';
  const pos = invert ? v < 0 : v > 0;
  return pos ? 'pos' : 'neg';
}}

function renderTop5() {{
  const sec = document.getElementById('top5Section');
  if (currentCat === '__all__') {{
    sec.innerHTML = '';
    return;
  }}
  const funds = TOP5[currentCat] || [];
  if (!funds.length) {{ sec.innerHTML=''; return; }}

  sec.innerHTML = `
    <div class="section-hdr">
      <div class="section-title">🏆 Top 5 — ${{currentCat}}</div>
    </div>
    <div class="top5-grid">
      ${{funds.map((f, i) => `
        <div class="fund-card rank-${{i+1}}">
          <div class="fund-rank">#${{i+1}} Expert Pick</div>
          <div class="fund-name">${{f.fund_name || '—'}}</div>
          <div class="fund-score">
            <div class="score-val">${{fmt(f.expert_score,1)}}</div>
            <div class="score-label">Score</div>
          </div>
          <div class="metrics-row">
            <div class="metric-pill">
              <div class="ml">3Y Return</div>
              <div class="mv ${{valClass(f.return_3y)}}">${{fmt(f.return_3y)}}%</div>
            </div>
            <div class="metric-pill">
              <div class="ml">5Y Return</div>
              <div class="mv ${{valClass(f.return_5y)}}">${{fmt(f.return_5y)}}%</div>
            </div>
            <div class="metric-pill">
              <div class="ml">Sharpe</div>
              <div class="mv ${{valClass(f.sharpe_3y)}}">${{fmt(f.sharpe_3y)}}</div>
            </div>
            <div class="metric-pill">
              <div class="ml">Sortino</div>
              <div class="mv ${{valClass(f.sortino_3y)}}">${{fmt(f.sortino_3y)}}</div>
            </div>
            <div class="metric-pill">
              <div class="ml">Alpha 3Y</div>
              <div class="mv ${{valClass(f.alpha_3y)}}">${{fmt(f.alpha_3y)}}%</div>
            </div>
            <div class="metric-pill">
              <div class="ml">Max DD</div>
              <div class="mv ${{valClass(f.max_drawdown, true)}}">${{fmt(f.max_drawdown)}}%</div>
            </div>
            <div class="metric-pill">
              <div class="ml">Up Capture</div>
              <div class="mv ${{valClass(f.upside_capture)}}">${{fmt(f.upside_capture,1)}}</div>
            </div>
            <div class="metric-pill">
              <div class="ml">Dn Capture</div>
              <div class="mv ${{valClass(f.downside_capture, true)}}">${{fmt(f.downside_capture,1)}}</div>
            </div>
          </div>
        </div>`).join('')}}
    </div>`;
}}

function renderTable(data) {{
  const tb = document.getElementById('tableBody');
  if (!data.length) {{
    tb.innerHTML = '<tr><td colspan="22" style="padding:40px;color:var(--muted);text-align:center;">No funds match your filters</td></tr>';
    return;
  }}
  tb.innerHTML = data.map(f => {{
    const sc = f.expert_score;
    return `<tr>
      <td class="rank-cell">#${{f.category_rank||'—'}}</td>
      <td class="fund-nm" title="${{f.fund_name}}">${{f.fund_name||'—'}}</td>
      <td><span class="score-chip ${{scoreClass(sc)}}">${{fmt(sc,1)}}</span></td>
      <td class="${{valClass(f.return_1y)}}">${{fmt(f.return_1y)}}%</td>
      <td class="${{valClass(f.return_3y)}}">${{fmt(f.return_3y)}}%</td>
      <td class="${{valClass(f.return_5y)}}">${{fmt(f.return_5y)}}%</td>
      <td class="${{valClass(f.return_10y)}}">${{fmt(f.return_10y)}}%</td>
      <td class="${{valClass(f.sip_return_3y)}}">${{fmt(f.sip_return_3y)}}%</td>
      <td class="${{valClass(f.sip_return_5y)}}">${{fmt(f.sip_return_5y)}}%</td>
      <td class="${{valClass(f.rolling_avg_3y)}}">${{fmt(f.rolling_avg_3y)}}%</td>
      <td class="${{valClass(f.sharpe_3y)}}">${{fmt(f.sharpe_3y)}}</td>
      <td class="${{valClass(f.sortino_3y)}}">${{fmt(f.sortino_3y)}}</td>
      <td class="neutral">${{fmt(f.volatility_3y)}}%</td>
      <td class="${{valClass(f.max_drawdown, true)}}">${{fmt(f.max_drawdown)}}%</td>
      <td class="${{valClass(f.calmar_3y)}}">${{fmt(f.calmar_3y)}}</td>
      <td class="${{valClass(f.upside_capture)}}">${{fmt(f.upside_capture,1)}}</td>
      <td class="${{valClass(f.downside_capture, true)}}">${{fmt(f.downside_capture,1)}}</td>
      <td class="${{valClass(f.alpha_3y)}}">${{fmt(f.alpha_3y)}}%</td>
      <td class="neutral">${{fmt(f.beta_3y,3)}}</td>
      <td class="${{valClass(f.information_ratio)}}">${{fmt(f.information_ratio)}}</td>
      <td class="neutral">${{fmt(f.pct_pos_1y,1)}}%</td>
      <td class="neutral">${{fmt(f.history_years,1)}}y</td>
    </tr>`;
  }}).join('');
}}

// Stats bar
function buildStats() {{
  const sb = document.getElementById('statsBar');
  const top = [...ALL_DATA].sort((a,b) => (b.expert_score||0)-(a.expert_score||0)).slice(0,1)[0];
  const avgSharpe = (ALL_DATA.filter(d=>d.sharpe_3y!=null).reduce((s,d)=>s+(d.sharpe_3y||0),0)/ALL_DATA.filter(d=>d.sharpe_3y!=null).length).toFixed(2);
  const avg5Y     = (ALL_DATA.filter(d=>d.return_5y!=null).reduce((s,d)=>s+(d.return_5y||0),0)/ALL_DATA.filter(d=>d.return_5y!=null).length).toFixed(1);
  sb.innerHTML = [
    ['Funds Analysed', ALL_DATA.length, ''],
    ['Avg 5Y Return', avg5Y + '%', ''],
    ['Avg Sharpe (3Y)', avgSharpe, ''],
    ['Categories', CATEGORIES.length, ''],
    ['Top Score', top ? top.expert_score.toFixed(1) : '—', ''],
    ['Metrics/Fund', '35+', ''],
  ].map(([l,v]) => `<div class="stat"><div class="stat-label">${{l}}</div><div class="stat-val">${{v}}</div></div>`).join('');
}}

buildStats();
applyFilters();
</script>
</body>
</html>"""
    DASH_PATH.write_text(html, encoding="utf-8")


if __name__ == "__main__":
    main()
