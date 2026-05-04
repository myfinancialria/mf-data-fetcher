"""
STEP 2 — COMPUTE ALL METRICS & RATIOS FOR COMPLETE UNIVERSE
Input  : data/nav_history/*.csv
         data/all_schemes.csv
Output : data/mf_complete_metrics.xlsx
         data/mf_complete_metrics.csv
"""

import pandas as pd
import numpy as np
import os
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR       = "data"
NAV_DIR        = os.path.join(DATA_DIR, "nav_history")
SCHEMES_FILE   = os.path.join(DATA_DIR, "all_schemes.csv")
OUTPUT_XLSX    = os.path.join(DATA_DIR, "mf_complete_metrics.xlsx")
OUTPUT_CSV     = os.path.join(DATA_DIR, "mf_complete_metrics.csv")
MAX_WORKERS    = 8
RISK_FREE      = 0.065 / 252   # daily risk-free rate

logging.basicConfig(level=logging.WARNING, format="%(asctime)s %(message)s")
log = logging.getLogger(__name__)


def load_nav(scheme_code):
    path = os.path.join(NAV_DIR, f"{scheme_code}.csv")
    if not os.path.exists(path):
        return None
    try:
        df = pd.read_csv(path, parse_dates=['date'])
        df = df.dropna().sort_values('date')
        nav = df.set_index('date')['nav'].astype(float)
        return nav
    except Exception:
        return None


def load_benchmark():
    for code in ['120716', '120503', '101835', '118989']:
        nav = load_nav(code)
        if nav is not None and len(nav) > 500:
            return nav
    return None


def safe_percentile(arr, q):
    try:
        if arr is None or len(arr) == 0:
            return None
        result = np.percentile(arr, q)
        return float(result) if np.isfinite(result) else None
    except Exception:
        return None


def safe_float(val):
    try:
        f = float(val)
        return f if np.isfinite(f) else None
    except Exception:
        return None


def cagr_calc(nav, days):
    try:
        if len(nav) < days:
            return None
        today = nav.index[-1]
        past_date = today - timedelta(days=days)
        past_nav = nav[nav.index <= past_date]
        if len(past_nav) == 0:
            return None
        start = float(past_nav.iloc[-1])
        end = float(nav.iloc[-1])
        if start <= 0:
            return None
        years = days / 365.25
        if days <= 365:
            return round((end / start - 1) * 100, 2)
        return round(((end / start) ** (1 / years) - 1) * 100, 2)
    except Exception:
        return None


def sip_return(nav, years):
    try:
        end_date = nav.index[-1]
        start_date = end_date - pd.DateOffset(years=years)
        dates = pd.date_range(start=start_date, end=end_date, freq='MS')
        if len(dates) < years * 10:
            return None
        total_units = 0.0
        total_invested = 0.0
        for d in dates:
            idx = nav.index.searchsorted(d)
            idx = min(idx, len(nav) - 1)
            price = float(nav.iloc[idx])
            if price > 0:
                total_units += 1000.0 / price
                total_invested += 1000.0
        if total_invested <= 0:
            return None
        current_value = total_units * float(nav.iloc[-1])
        n_years = (end_date - dates[0]).days / 365.25
        if n_years <= 0:
            return None
        ratio = current_value / total_invested
        return round((ratio ** (1 / n_years) - 1) * 100, 2)
    except Exception:
        return None


def compute_all_metrics(scheme_code, scheme_name, benchmark_nav):
    try:
        nav = load_nav(scheme_code)
        if nav is None or len(nav) < 252:
            return None

        returns = nav.pct_change().dropna()
        if len(returns) < 50:
            return None

        metrics = {
            'Scheme Code': scheme_code,
            'Scheme Name': scheme_name,
            'Latest NAV':  round(float(nav.iloc[-1]), 4),
            'As On Date':  nav.index[-1].strftime('%Y-%m-%d'),
        }

        # Returns
        for label, days in [('1W',7),('1M',30),('3M',91),('6M',182),
                             ('1Y',365),('2Y',730),('3Y',1095),
                             ('5Y',1825),('7Y',2555),('10Y',3650)]:
            metrics[f'Return_{label} (%)'] = cagr_calc(nav, days)

        # SIP Returns
        metrics['SIP_1Y (%)'] = sip_return(nav, 1)
        metrics['SIP_3Y (%)'] = sip_return(nav, 3)
        metrics['SIP_5Y (%)'] = sip_return(nav, 5)

        # Risk
        std_dev = safe_float(returns.std() * np.sqrt(252))
        neg_ret = returns[returns < 0]
        downside_dev = safe_float(neg_ret.std() * np.sqrt(252)) if len(neg_ret) > 5 else std_dev

        rolling_max = nav.cummax()
        drawdown = (nav - rolling_max) / rolling_max
        max_dd = safe_float(drawdown.min() * 100)

        # VaR — safe version
        ret_arr = returns.dropna().values
        var_95  = safe_percentile(ret_arr, 5)
        var_95  = round(var_95 * 100, 2) if var_95 is not None else None
        var_99  = safe_percentile(ret_arr, 1)
        var_99  = round(var_99 * 100, 2) if var_99 is not None else None

        # CVaR
        try:
            threshold = np.percentile(ret_arr, 5)
            tail = ret_arr[ret_arr <= threshold]
            cvar_95 = round(float(tail.mean()) * 100, 2) if len(tail) > 0 else None
        except Exception:
            cvar_95 = None

        metrics.update({
            'Std_Dev_Annual (%)': round(std_dev * 100, 2) if std_dev else None,
            'Downside_Dev (%)':   round(downside_dev * 100, 2) if downside_dev else None,
            'Max_Drawdown (%)':   max_dd,
            'VaR_95 (%)':         var_95,
            'VaR_99 (%)':         var_99,
            'CVaR_95 (%)':        cvar_95,
        })

        # Sharpe & Sortino
        excess = returns.mean() - RISK_FREE
        sharpe  = safe_float(excess / returns.std() * np.sqrt(252)) if returns.std() > 0 else None
        sortino = safe_float(excess / downside_dev * np.sqrt(252)) if (downside_dev and downside_dev > 0) else None
        cagr_3y = metrics.get('Return_3Y (%)')
        calmar  = round(cagr_3y / abs(max_dd), 3) if (cagr_3y and max_dd and max_dd != 0) else None

        metrics.update({
            'Sharpe_Ratio':  round(sharpe, 3)  if sharpe  else None,
            'Sortino_Ratio': round(sortino, 3) if sortino else None,
            'Calmar_Ratio':  calmar,
        })

        # Benchmark ratios
        if benchmark_nav is not None:
            try:
                bm_ret = benchmark_nav.pct_change().dropna()
                common = returns.index.intersection(bm_ret.index)
                if len(common) > 252:
                    fr = returns.reindex(common).dropna()
                    br = bm_ret.reindex(fr.index).dropna()
                    fr = fr.reindex(br.index)

                    cov = np.cov(fr, br)
                    beta = safe_float(cov[0][1] / cov[1][1]) if cov[1][1] != 0 else None

                    if beta:
                        alpha_d = fr.mean() - (RISK_FREE + beta * (br.mean() - RISK_FREE))
                        alpha_a = round(safe_float(alpha_d * 252 * 100), 2) if alpha_d else None
                    else:
                        alpha_a = None

                    treynor = safe_float(excess / beta * np.sqrt(252)) if beta else None
                    ar = fr - br
                    info = safe_float(ar.mean() / ar.std() * np.sqrt(252)) if ar.std() > 0 else None

                    up   = br > 0
                    dn   = br < 0
                    up_c = safe_float(fr[up].mean() / br[up].mean() * 100) if br[up].mean() != 0 else None
                    dn_c = safe_float(fr[dn].mean() / br[dn].mean() * 100) if br[dn].mean() != 0 else None
                    r2   = safe_float(np.corrcoef(fr, br)[0][1] ** 2)

                    metrics.update({
                        'Beta':                 round(beta, 3)    if beta    else None,
                        'Alpha_Annual (%)':     alpha_a,
                        'Treynor_Ratio':        round(treynor, 3) if treynor else None,
                        'Info_Ratio':           round(info, 3)    if info    else None,
                        'Upside_Capture (%)':   round(up_c, 1)    if up_c    else None,
                        'Downside_Capture (%)': round(dn_c, 1)    if dn_c    else None,
                        'R_Squared':            round(r2, 3)      if r2      else None,
                    })
            except Exception:
                pass

        # Consistency Score
        try:
            rolling_1y = returns.rolling(252).sum()
            consistency = round(float((rolling_1y > 0).mean() * 100), 1)
            metrics['Consistency_Score'] = consistency
        except Exception:
            metrics['Consistency_Score'] = None

        # Composite Score
        score = 0.0
        weights = {
            'Return_3Y (%)': 0.20, 'Return_5Y (%)': 0.15,
            'Sharpe_Ratio': 0.20,  'Sortino_Ratio': 0.10,
            'Alpha_Annual (%)': 0.15, 'Calmar_Ratio': 0.10,
            'Consistency_Score': 0.10
        }
        for k, w in weights.items():
            v = metrics.get(k)
            if v is not None:
                score += float(v) * w
        metrics['Composite_Score'] = round(score, 2)

        return metrics

    except Exception as e:
        log.debug(f"Failed {scheme_code}: {e}")
        return None


def compute_metrics_for_all(schemes_df):
    print("📈 Loading benchmark...")
    benchmark_nav = load_benchmark()
    if benchmark_nav is not None:
        print(f"   Benchmark loaded ({len(benchmark_nav)} days)")
    else:
        print("   ⚠ No benchmark — Beta/Alpha skipped")

    nav_files = set(f.replace('.csv', '') for f in os.listdir(NAV_DIR))
    available = schemes_df[schemes_df['scheme_code'].astype(str).isin(nav_files)].copy()
    print(f"\n📊 Processing {len(available):,} funds with downloaded NAV data...\n")

    results = []

    def _compute(row):
        return compute_all_metrics(str(row['scheme_code']), str(row['scheme_name']), benchmark_nav)

    rows = available.to_dict('records')

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_compute, row): row for row in rows}
        with tqdm(total=len(rows), desc="⚙ Computing metrics", unit="fund") as pbar:
            for future in as_completed(futures):
                try:
                    result = future.result()
                    if result:
                        row = futures[future]
                        result['Category'] = row.get('category', 'Other')
                        results.append(result)
                except Exception:
                    pass
                pbar.update(1)

    df = pd.DataFrame(results)
    df = df.sort_values('Composite_Score', ascending=False).reset_index(drop=True)
    df['Rank'] = df.index + 1
    print(f"\n✅ Metrics computed for {len(df):,} funds")
    return df


def build_excel(df, path):
    print(f"\n📊 Building Excel: {path}")
    wb = Workbook()

    # Sheet 1 — All Funds
    ws = wb.active
    ws.title = "All Funds"
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(color="FFFFFF", bold=True, size=9)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = Alignment(horizontal="center")
            c.font = Font(size=9)
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="EEF3FA")

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(len(str(h)) + 3, 24)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Sheet 2 — Top 10 Per Category
    ws2 = wb.create_sheet("Top Per Category")
    cols = ['Rank','Scheme Name','Category','Return_1Y (%)','Return_3Y (%)','Return_5Y (%)',
            'Sharpe_Ratio','Sortino_Ratio','Beta','Alpha_Annual (%)','Max_Drawdown (%)',
            'Upside_Capture (%)','Downside_Capture (%)','Composite_Score']
    cols = [c for c in cols if c in df.columns]

    row_num = 1
    COLORS = {"Large Cap":"2E75B6","Mid Cap":"ED7D31","Small Cap":"70AD47",
               "Flexi Cap":"7030A0","ELSS":"C00000","Index Fund":"4472C4"}

    for cat in sorted(df['Category'].unique()):
        cat_df = df[df['Category'] == cat].nlargest(10, 'Composite_Score')
        if len(cat_df) == 0:
            continue
        color = COLORS.get(cat, "1F4E79")
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(cols))
        hc = ws2.cell(row=row_num, column=1, value=f"🏆  {cat.upper()} — TOP {len(cat_df)}")
        hc.fill = PatternFill("solid", fgColor=color)
        hc.font = Font(color="FFFFFF", bold=True, size=11)
        hc.alignment = Alignment(horizontal="center")
        row_num += 1
        for ci, h in enumerate(cols, 1):
            c = ws2.cell(row=row_num, column=ci, value=h)
            c.fill = PatternFill("solid", fgColor="D6E4F0")
            c.font = Font(bold=True, size=9)
            c.alignment = Alignment(horizontal="center")
        row_num += 1
        for rk, (_, r) in enumerate(cat_df.iterrows(), 1):
            for ci, h in enumerate(cols, 1):
                val = rk if h == 'Rank' else r.get(h, '')
                c = ws2.cell(row=row_num, column=ci, value=val)
                c.alignment = Alignment(horizontal="center")
                c.font = Font(size=9, bold=(rk == 1), color="155724" if rk == 1 else "000000")
                if rk == 1:
                    c.fill = PatternFill("solid", fgColor="D1ECE1")
            row_num += 1
        row_num += 1

    for ci, h in enumerate(cols, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = min(len(h) + 4, 28)
    ws2.freeze_panes = "A1"

    wb.save(path)
    print(f"✅ Saved: {path}")


def main():
    print("=" * 65)
    print("  ⚙  MUTUAL FUND COMPLETE METRICS COMPUTATION")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65)

    if not os.path.exists(SCHEMES_FILE):
        print(f"❌ {SCHEMES_FILE} not found. Run fetch_universe.py first!")
        return

    schemes_df = pd.read_csv(SCHEMES_FILE, dtype={'scheme_code': str})
    print(f"📋 Loaded {len(schemes_df):,} schemes")

    df = compute_metrics_for_all(schemes_df)
    df.to_csv(OUTPUT_CSV, index=False)
    print(f"💾 CSV: {OUTPUT_CSV}")

    build_excel(df, OUTPUT_XLSX)

    print(f"\n{'='*65}")
    print(f"  ✅ DONE! {len(df):,} funds | 35+ metrics each")
    print(f"  📊 Open: {OUTPUT_XLSX}")
    print(f"{'='*65}")

    top_cols = [c for c in ['Rank','Scheme Name','Category','Return_3Y (%)','Sharpe_Ratio','Composite_Score'] if c in df.columns]
    print("\nTop 5 funds:")
    print(df[top_cols].head(5).to_string(index=False))


if __name__ == "__main__":
    main()
