"""
╔══════════════════════════════════════════════════════════════════╗
║     HNI / UHNI GOALS-BASED WEALTH MANAGEMENT ENGINE  v2         ║
║     ✅ 1-2 funds per category only (no confusion)               ║
║     ✅ SIP amount shown per fund                                 ║
║     ✅ Same fund NEVER repeated across goals                     ║
║     ✅ Clear allocation % per fund                               ║
╚══════════════════════════════════════════════════════════════════╝
"""

import pandas as pd
import numpy as np
import os
import re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
#  ✏  CLIENT PROFILE  —  EDIT THIS
# ══════════════════════════════════════════════════════════════════

CLIENT_PROFILE = {
    "name":            "Nithin",
    "age":             35,
    "annual_income":   3_000_000,
    "monthly_surplus": 150_000,
    "existing_corpus": 5_000_000,
    "tax_bracket":     30,
    "risk_profile":    "ModerateAggressive",
    # Options: Conservative | ModerateConservative | Moderate
    #          ModerateAggressive | Aggressive | UltraAggressive

    # OPTIONAL ASSETS — set True only when client specifically requests
    # Gold is ALWAYS included automatically (mandatory per HNI framework)
    "alternatives": {
        "reits":         False,   # REIT funds — passive real estate income (opt-in)
        "international": False,   # US/Global equity — currency diversification (opt-in)
    }
}

# ══════════════════════════════════════════════════════════════════
#  ✏  YOUR GOALS  —  EDIT THIS
# ══════════════════════════════════════════════════════════════════

GOALS = [
    {
        "name":          "Retirement Corpus",
        "goal_type":     "Retirement",
        "target_amount": 100_000_000,
        "time_years":    25,
        "monthly_sip":   50_000,
        "lumpsum":       500_000,
        "priority":      "MustHave",
    },
    {
        "name":          "Child Higher Education",
        "goal_type":     "ChildEducation",
        "target_amount": 10_000_000,
        "time_years":    15,
        "monthly_sip":   20_000,
        "lumpsum":       0,
        "priority":      "MustHave",
    },
    {
        "name":          "ELSS Tax Saving",
        "goal_type":     "TaxSaving",
        "target_amount": 500_000,
        "time_years":    3,
        "monthly_sip":   12_500,
        "lumpsum":       0,
        "priority":      "MustHave",
    },
    {
        "name":          "Dream Home Down Payment",
        "goal_type":     "HomePurchase",
        "target_amount": 7_500_000,
        "time_years":    7,
        "monthly_sip":   40_000,
        "lumpsum":       1_000_000,
        "priority":      "ShouldHave",
    },
    {
        "name":          "Wealth Creation Legacy",
        "goal_type":     "WealthCreation",
        "target_amount": 200_000_000,
        "time_years":    20,
        "monthly_sip":   30_000,
        "lumpsum":       2_000_000,
        "priority":      "Aspirational",
    },
]

# ══════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════

DATA_DIR    = "data"
METRICS_CSV = os.path.join(DATA_DIR, "mf_complete_metrics.csv")
OUTPUT_FILE = os.path.join(DATA_DIR, "HNI_Goal_Portfolio.xlsx")

EXPECTED_EQUITY = 0.13
EXPECTED_DEBT   = 0.075
RISK_FREE       = 0.065 / 252

GOAL_ICONS = {
    "Retirement": "R", "ChildEducation": "E", "ChildMarriage": "M",
    "HomePurchase": "H", "WealthCreation": "W", "TaxSaving": "T",
    "EmergencyFund": "S", "LuxuryPurchase": "L", "WealthPreservation": "P",
}

PRIORITY_COLORS = {
    "MustHave":    ("C00000", "FFE7E7", "FF4444"),
    "ShouldHave":  ("ED7D31", "FFF3E7", "FF8C42"),
    "Aspirational":("7030A0", "F3E7FF", "9B59B6"),
}

# ══════════════════════════════════════════════════════════════════
#  ALLOCATION MATRIX
# ══════════════════════════════════════════════════════════════════

ALLOCATION_MATRIX = {
    "Conservative":        {(0,1):(0,90,10),(1,3):(20,75,5),(3,5):(40,55,5),(5,7):(55,40,5),(7,10):(65,32,3),(10,99):(70,28,2)},
    "ModerateConservative":{(0,1):(5,90,5),(1,3):(30,65,5),(3,5):(50,45,5),(5,7):(65,32,3),(7,10):(75,23,2),(10,99):(80,18,2)},
    "Moderate":            {(0,1):(10,85,5),(1,3):(40,55,5),(3,5):(60,37,3),(5,7):(75,23,2),(7,10):(82,16,2),(10,99):(87,11,2)},
    "ModerateAggressive":  {(0,1):(20,75,5),(1,3):(55,42,3),(3,5):(72,26,2),(5,7):(83,15,2),(7,10):(88,10,2),(10,99):(92,6,2)},
    "Aggressive":          {(0,1):(30,65,5),(1,3):(65,32,3),(3,5):(80,18,2),(5,7):(90,8,2),(7,10):(93,5,2),(10,99):(95,3,2)},
    "UltraAggressive":     {(0,1):(40,55,5),(1,3):(75,23,2),(3,5):(88,10,2),(5,7):(95,3,2),(7,10):(97,1,2),(10,99):(98,0,2)},
}

# Equity sub-allocation by time horizon (% of equity bucket)
EQUITY_SUB = {
    (0,3):   [("Large Cap",70),("Flexi Cap",30)],
    (3,5):   [("Large Cap",50),("Flexi Cap",30),("Mid Cap",20)],
    (5,7):   [("Large Cap",40),("Flexi Cap",30),("Mid Cap",30)],
    (7,10):  [("Large Cap",35),("Flexi Cap",30),("Mid Cap",25),("Small Cap",10)],
    (10,99): [("Large Cap",30),("Flexi Cap",25),("Mid Cap",25),("Small Cap",20)],
}

DEBT_SUB = {
    (0,1):   [("Liquid Fund",100)],
    (1,3):   [("Short Duration",60),("Corporate Bond",40)],
    (3,5):   [("Corporate Bond",50),("Banking & PSU",50)],
    (5,99):  [("Dynamic Bond",50),("Corporate Bond",50)],
}

# ── GOLD — mandatory in ALL portfolios (inflation hedge) ──────────────
# % taken from equity allocation. Varies by risk profile.
GOLD_ALLOCATION = {
    "Conservative":         10,   # 10% gold — capital preservation focus
    "ModerateConservative":  9,
    "Moderate":              8,
    "ModerateAggressive":    7,
    "Aggressive":            6,
    "UltraAggressive":       5,   # 5% gold — growth focus but still hedged
}
GOLD_MIN_YEARS   = 3              # Gold not added for goals < 3 years
GOLD_CATEGORIES  = ["Gold Fund", "ETF - Equity"]  # try in order

# ── OPTIONAL ASSETS — added only when client opts in ──────────────────
# % of EQUITY reallocated to each
ALTERNATIVES_CONFIG = {
    "reits":         {"pct_of_equity": 7,  "category": "FOF - Domestic",    "bucket": "REITs"},
    "international": {"pct_of_equity": 10, "category": "International FOF", "bucket": "International"},
}
ALTERNATIVES_MIN_YEARS = {"reits": 5, "international": 5}

GOAL_CAT_OVERRIDE = {
    # (equity_override, debt_override)
    # None = use standard allocation for that bucket
    # []   = skip that bucket entirely (e.g. EmergencyFund has no equity)
    "TaxSaving":          (["ELSS"],  None),           # ELSS for equity, normal debt funds
    "EmergencyFund":      ([],        ["Liquid Fund"]), # Liquid only, no equity
    "WealthPreservation": (["Large Cap","Balanced Advantage"], ["Corporate Bond"]),
}

HNI_FILTERS = {
    "Conservative":        {"min_sharpe":0.3,  "max_dd":-20,"min_con":60},
    "ModerateConservative":{"min_sharpe":0.35, "max_dd":-25,"min_con":62},
    "Moderate":            {"min_sharpe":0.4,  "max_dd":-30,"min_con":60},
    "ModerateAggressive":  {"min_sharpe":0.45, "max_dd":-38,"min_con":58},
    "Aggressive":          {"min_sharpe":0.5,  "max_dd":-45,"min_con":55},
    "UltraAggressive":     {"min_sharpe":0.3,  "max_dd":-60,"min_con":50},
}


# ══════════════════════════════════════════════════════════════════
#  CORE LOGIC
# ══════════════════════════════════════════════════════════════════

def time_bucket(yrs):
    for lo,hi in [(0,1),(1,3),(3,5),(5,7),(7,10),(10,99)]:
        if lo <= yrs < hi:
            return (lo,hi)
    return (10,99)


def get_allocation(risk, yrs):
    matrix = ALLOCATION_MATRIX.get(risk, ALLOCATION_MATRIX["Moderate"])
    return matrix.get(time_bucket(yrs),(87,11,2))


def project_corpus(sip, lump, yrs, eq_pct, debt_pct):
    rate  = (eq_pct/100*EXPECTED_EQUITY) + (debt_pct/100*EXPECTED_DEBT)
    mr    = rate/12
    m     = yrs*12
    fv_l  = lump*((1+rate)**yrs)
    fv_s  = sip*(((1+mr)**m-1)/mr*(1+mr)) if mr>0 else sip*m
    return round(fv_l+fv_s)


def req_sip(target, lump, yrs, eq_pct, debt_pct):
    rate  = (eq_pct/100*EXPECTED_EQUITY)+(debt_pct/100*EXPECTED_DEBT)
    mr    = rate/12
    m     = yrs*12
    rem   = target - lump*((1+rate)**yrs)
    if rem <= 0: return 0
    return max(0,round(rem/((((1+mr)**m-1)/mr*(1+mr)) if mr>0 else m)))


def pick_funds(df, category, risk, used_codes, n=2):
    """
    Pick top N funds from a category.
    NEVER picks a fund code already used in another goal.
    """
    filt   = HNI_FILTERS.get(risk, HNI_FILTERS["Moderate"])
    cat_df = df[df['Category']==category].copy()

    if len(cat_df)==0:
        cat_df = df[df['Category'].str.contains(
            category.split()[0],case=False,na=False)].copy()
    if len(cat_df)==0:
        return pd.DataFrame()

    # Exclude already-used funds across other goals
    cat_df = cat_df[~cat_df['Scheme Code'].astype(str).isin(used_codes)]

    # Apply quality filters
    f = cat_df.copy()
    if 'Sharpe_Ratio' in f.columns:
        f = f[f['Sharpe_Ratio'].fillna(0) >= filt['min_sharpe']]
    if 'Max_Drawdown (%)' in f.columns:
        f = f[f['Max_Drawdown (%)'].fillna(-999) >= filt['max_dd']]
    if 'Consistency_Score' in f.columns:
        f = f[f['Consistency_Score'].fillna(0) >= filt['min_con']]

    # Fallback: relax filters if too few
    if len(f) < n:
        f = cat_df.copy()

    if 'Composite_Score' in f.columns:
        f = f.sort_values('Composite_Score', ascending=False)

    return f.head(n)


def build_portfolio(goal, df, risk, used_codes):
    yrs                     = goal['time_years']
    eq_pct,debt_pct,liq_pct = get_allocation(risk, yrs)
    tb                      = time_bucket(yrs)
    monthly_sip             = goal['monthly_sip']

    # ── Category lists ────────────────────────────────────────
    default_eq_cats   = [c for c,_ in EQUITY_SUB.get(tb, EQUITY_SUB[(10,99)])]
    default_debt_cats = [c for c,_ in DEBT_SUB.get(tb, DEBT_SUB[(5,99)])]

    if goal['goal_type'] in GOAL_CAT_OVERRIDE:
        ov_eq, ov_debt = GOAL_CAT_OVERRIDE[goal['goal_type']]
        eq_cats   = ov_eq   if ov_eq   is not None else default_eq_cats
        debt_cats = ov_debt if ov_debt is not None else default_debt_cats
    else:
        eq_cats   = default_eq_cats
        debt_cats = default_debt_cats

    # ── Weights ───────────────────────────────────────────────
    std_eq_w   = {c:w for c,w in EQUITY_SUB.get(tb, EQUITY_SUB[(10,99)])}
    std_debt_w = {c:w for c,w in DEBT_SUB.get(tb, DEBT_SUB[(5,99)])}

    # When categories are overridden, split equally among them
    if goal['goal_type'] in GOAL_CAT_OVERRIDE and GOAL_CAT_OVERRIDE[goal['goal_type']][0] is not None:
        n = len(eq_cats)
        eq_weights = {c: round(100/n) for c in eq_cats} if n > 0 else std_eq_w
    else:
        eq_weights = std_eq_w

    if goal['goal_type'] in GOAL_CAT_OVERRIDE and GOAL_CAT_OVERRIDE[goal['goal_type']][1] is not None:
        n2 = len(debt_cats)
        debt_weights = {c: round(100/n2) for c in debt_cats} if n2 > 0 else std_debt_w
    else:
        debt_weights = std_debt_w

    # ── Alternatives opted in by client ───────────────────────
    alt_config  = CLIENT_PROFILE.get('alternatives', {})
    active_alts = []
    total_alt_pct = 0

    # GOLD — always mandatory for goals >= GOLD_MIN_YEARS
    gold_pct = 0
    if yrs >= GOLD_MIN_YEARS and goal['goal_type'] not in ('EmergencyFund', 'TaxSaving'):
        gold_pct = GOLD_ALLOCATION.get(risk, 7)
        total_alt_pct += gold_pct

    # OPTIONAL — REITs and International only when client opts in
    for ak, cfg in ALTERNATIVES_CONFIG.items():
        if alt_config.get(ak, False) and yrs >= ALTERNATIVES_MIN_YEARS.get(ak, 5):
            active_alts.append((ak, cfg))
            total_alt_pct += cfg['pct_of_equity']

    actual_eq_pct = max(0, eq_pct - total_alt_pct)

    fund_plan = []

    def add_fund(f, bucket, cat, alloc, sip):
        code = str(f.get('Scheme Code',''))
        used_codes.add(code)
        fund_plan.append({
            "bucket":       bucket, "category":    cat,
            "scheme_code":  code,   "scheme_name": f.get('Scheme Name','—'),
            "alloc_pct":    alloc,  "monthly_sip": sip,
            "return_1y":    f.get('Return_1Y (%)'),
            "return_3y":    f.get('Return_3Y (%)'),
            "return_5y":    f.get('Return_5Y (%)'),
            "sharpe":       f.get('Sharpe_Ratio'),
            "sortino":      f.get('Sortino_Ratio') if bucket=="Equity" else None,
            "alpha":        f.get('Alpha_Annual (%)') if bucket=="Equity" else None,
            "max_dd":       f.get('Max_Drawdown (%)'),
            "upside_cap":   f.get('Upside_Capture (%)') if bucket=="Equity" else None,
            "downside_cap": f.get('Downside_Capture (%)') if bucket=="Equity" else None,
            "composite":    f.get('Composite_Score'),
        })

    # ── EQUITY ────────────────────────────────────────────────
    equity_sip = monthly_sip * actual_eq_pct / 100
    for cat in eq_cats:
        w = eq_weights.get(cat, 0)
        if w == 0: continue
        funds = pick_funds(df, cat, risk, used_codes, n=1)
        if len(funds) == 0: continue
        per_sip   = round(equity_sip * w/100 / len(funds) / 500) * 500
        per_alloc = round(actual_eq_pct * w/100 / len(funds), 1)
        for _, f in funds.iterrows():
            add_fund(f, "Equity", cat, per_alloc, per_sip)

    # ── DEBT ──────────────────────────────────────────────────
    debt_sip = monthly_sip * debt_pct / 100
    for cat in debt_cats:
        w = debt_weights.get(cat, 0)
        if w == 0: continue
        funds = pick_funds(df, cat, risk, used_codes, n=1)
        if len(funds) == 0: continue
        per_sip   = round(debt_sip * w/100 / len(funds) / 500) * 500
        per_alloc = round(debt_pct * w/100 / len(funds), 1)
        for _, f in funds.iterrows():
            add_fund(f, "Debt", cat, per_alloc, per_sip)

    # ── GOLD (mandatory) ─────────────────────────────────────
    if gold_pct > 0:
        gold_sip   = round(monthly_sip * gold_pct / 100 / 500) * 500
        for gold_cat in GOLD_CATEGORIES:
            funds = pick_funds(df, gold_cat, risk, used_codes, n=1)
            if len(funds) > 0:
                for _, f in funds.iterrows():
                    add_fund(f, "Gold", gold_cat, gold_pct, gold_sip)
                break   # stop after first category that has funds

    # ── OPTIONAL ALTERNATIVES (REITs / International) ────────
    for ak, cfg in active_alts:
        funds = pick_funds(df, cfg['category'], risk, used_codes, n=1)
        if len(funds) == 0: continue
        alt_sip   = round(monthly_sip * cfg['pct_of_equity']/100 / 500) * 500
        alt_alloc = cfg['pct_of_equity']
        for _, f in funds.iterrows():
            add_fund(f, cfg['bucket'], cfg['category'], alt_alloc, alt_sip)

    proj     = project_corpus(monthly_sip, goal['lumpsum'], yrs, eq_pct, debt_pct)
    req      = req_sip(goal['target_amount'], goal['lumpsum'], yrs, eq_pct, debt_pct)

    return {
        "goal":       goal,
        "eq_pct":     actual_eq_pct,
        "debt_pct":   debt_pct,
        "liq_pct":    liq_pct,
        "gold_pct":   gold_pct,
        "alt_pcts":   {k:v['pct_of_equity'] for k,v in active_alts},
        "fund_plan":  fund_plan,
        "projected":  proj,
        "req_sip":    req,
        "on_track":   proj >= goal['target_amount'],
        "gap":        goal['target_amount'] - proj,
    }


# ══════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════

def cr(v):
    try:
        v = float(v)
        if abs(v) >= 1e7: return f"₹{v/1e7:.2f} Cr"
        if abs(v) >= 1e5: return f"₹{v/1e5:.2f} L"
        return f"₹{v:,.0f}"
    except: return "—"

def fmt(v, suffix="", decimals=2):
    try:
        if v is None or (isinstance(v,float) and np.isnan(v)): return "—"
        return f"{float(v):.{decimals}f}{suffix}"
    except: return "—"

def bdr():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)

def H(ws, r, c, val, bg="1F4E79", fg="FFFFFF", bold=True, sz=9,
      al="center", wrap=False, colspan=1):
    if colspan > 1:
        ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+colspan-1)
    cell = ws.cell(row=r,column=c,value=val)
    cell.fill      = PatternFill("solid",fgColor=bg)
    cell.font      = Font(color=fg,bold=bold,size=sz,name="Arial")
    cell.alignment = Alignment(horizontal=al,vertical="center",wrap_text=wrap)
    cell.border    = bdr()
    return cell

def C(ws, r, c, val, bg=None, fg="000000", bold=False, sz=9,
      al="center", wrap=False):
    cell = ws.cell(row=r,column=c,value=val)
    if bg: cell.fill = PatternFill("solid",fgColor=bg)
    cell.font      = Font(color=fg,bold=bold,size=sz,name="Arial")
    cell.alignment = Alignment(horizontal=al,vertical="center",wrap_text=wrap)
    cell.border    = bdr()
    return cell


# ══════════════════════════════════════════════════════════════════
#  SHEET 1: DASHBOARD
# ══════════════════════════════════════════════════════════════════

def sheet_dashboard(wb, client, portfolios):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:M1")
    ws["A1"].value = f"  HNI GOALS-BASED PORTFOLIO  |  {client['name'].upper()}  |  Risk: {client['risk_profile']}  |  {date.today().strftime('%d %B %Y')}"
    ws["A1"].fill  = PatternFill("solid",fgColor="0D1B2A")
    ws["A1"].font  = Font(color="FFD700",bold=True,size=13,name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:M2")
    ws["A2"].value = f"  Age: {client['age']}  |  Monthly Surplus: {cr(client['monthly_surplus'])}  |  Existing Corpus: {cr(client['existing_corpus'])}  |  Tax Bracket: {client['tax_bracket']}%"
    ws["A2"].fill  = PatternFill("solid",fgColor="1A2E44")
    ws["A2"].font  = Font(color="A0B4C8",size=10,name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    # Goals summary table
    row = 4
    heads = ["Goal","Priority","Time","Target","Monthly SIP","Projected","Req. SIP",
             "Status","Eq %","Debt %","Gold %","No. of Funds","Sheet"]
    for ci,h in enumerate(heads,1):
        H(ws,row,ci,h,bg="1F4E79",sz=9)
    ws.row_dimensions[row].height = 22
    row += 1

    total_sip = 0
    for i,p in enumerate(portfolios,1):
        g    = p['goal']
        hc,bgc,_ = PRIORITY_COLORS.get(g['priority'],("333333","F5F5F5","999999"))
        st_bg = "D1ECE1" if p['on_track'] else "FFF3CD"
        st_fg = "155724" if p['on_track'] else "856404"
        status= "On Track" if p['on_track'] else f"Need {cr(p['req_sip'])}/mo"
        nfunds= len(p['fund_plan'])
        gold_p = p.get('gold_pct', 0)
        alloc_str = f"{p['eq_pct']}% Eq | {p['debt_pct']}% Debt"
        if gold_p: alloc_str += f" | {gold_p}% Gold"

        vals = [g['name'], g['priority'], f"{g['time_years']}Y",
                cr(g['target_amount']), cr(g['monthly_sip']),
                cr(p['projected']), cr(p['req_sip']),
                status, f"{p['eq_pct']}%", f"{p['debt_pct']}%",
                f"{gold_p}%" if gold_p else "—", nfunds, f"See sheet {i+1}"]

        for ci,v in enumerate(vals,1):
            c = C(ws,row,ci,v,bg=bgc,sz=9,al="left" if ci==1 else "center")
            if ci==2: c.font = Font(color=hc,bold=True,sz=9,name="Arial")
            if ci==8:
                c.fill = PatternFill("solid",fgColor=st_bg)
                c.font = Font(color=st_fg,bold=True,size=9,name="Arial")
        ws.row_dimensions[row].height = 20
        total_sip += g['monthly_sip']
        row += 1

    # Total row
    row += 1
    C(ws,row,1,"TOTAL MONTHLY SIP COMMITMENT",bg="0D1B2A",fg="FFFFFF",bold=True,al="left")
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    C(ws,row,5,cr(total_sip),bg="FFD700",fg="000000",bold=True,sz=11)
    surplus_pct = round(total_sip/client['monthly_surplus']*100)
    C(ws,row,6,f"{surplus_pct}% of monthly surplus",bg="FFD700",fg="555555",sz=9)
    ws.merge_cells(start_row=row,start_column=6,end_row=row,end_column=13)
    ws.row_dimensions[row].height = 26

    # Column widths
    for ci,w in enumerate([30,16,8,14,14,14,14,20,8,8,12,14],1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ══════════════════════════════════════════════════════════════════
#  SHEET 2+: ONE SHEET PER GOAL — FUND PLAN
# ══════════════════════════════════════════════════════════════════

def sheet_goal(wb, p, sheet_num):
    g       = p['goal']
    safe    = re.sub(r'[\/\\\?\*\[\]\:\']','',g['name'])[:28]
    ws      = wb.create_sheet(safe)
    ws.sheet_view.showGridLines = False
    hc,bgc,ac = PRIORITY_COLORS.get(g['priority'],("1F4E79","EEF5FF","4472C4"))

    # ── Header ────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"].value = f"  {g['name'].upper()}  |  {g['goal_type']}  |  {g['priority']}  |  Target: {cr(g['target_amount'])}  |  {g['time_years']} Years"
    ws["A1"].fill  = PatternFill("solid",fgColor="0D1B2A")
    ws["A1"].font  = Font(color="FFD700",bold=True,size=12,name="Arial")
    ws["A1"].alignment = Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Summary boxes ─────────────────────────────────────────
    boxes = [
        ("Target Corpus",    cr(g['target_amount'])),
        ("Time Horizon",     f"{g['time_years']} Years"),
        ("Your Monthly SIP", cr(g['monthly_sip'])),
        ("Lumpsum Today",    cr(g['lumpsum'])),
        ("Projected Corpus", cr(p['projected'])),
        ("Recommended SIP",  cr(p['req_sip'])),
        ("Asset Allocation", f"{p['eq_pct']}% Eq | {p['debt_pct']}% Debt"),
        ("Status",           "✅ On Track" if p['on_track'] else "⚠ Top-Up SIP"),
    ]
    for i,(label,value) in enumerate(boxes,1):
        H(ws,2,i,label,bg="2E4057",fg="A0B4C8",sz=8)
        is_status = (label=="Status")
        is_proj   = (label=="Projected Corpus")
        vbg = ("D1ECE1" if p['on_track'] else "FFF3CD") if is_status or is_proj else bgc
        C(ws,3,i,value,bg=vbg,bold=True,sz=10,
          fg=("155724" if p['on_track'] else "856404") if is_status else "000000")
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 10

    # ── Fund Plan table ───────────────────────────────────────
    row = 5
    ws.merge_cells(f"A{row}:L{row}")
    H(ws,row,1,f"  RECOMMENDED FUND ALLOCATION  —  {len(p['fund_plan'])} FUNDS  |  Total SIP: {cr(g['monthly_sip'])} / month",
      bg="1F4E79",sz=11,al="left",colspan=1)
    ws.merge_cells(f"A{row}:L{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    # Column headers
    cols = [
        ("Bucket",      8),  ("Category",    16), ("Fund Name",    36),
        ("Alloc %",     9),  ("Monthly SIP", 14), ("1Y Return",    10),
        ("3Y Return",   10), ("5Y Return",   10), ("Sharpe",        9),
        ("Alpha %",      9), ("Max Drawdown", 12), ("Why This Fund", 30),
    ]
    for ci,(h,w) in enumerate(cols,1):
        H(ws,row,ci,h,bg="2E4057",sz=9)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 20
    row += 1

    BUCKET_BG = {"Equity":"EBF5FB","Debt":"EAFAF1","Liquid":"FEF9E7"}
    BUCKET_HDR= {"Equity":"1B4F72","Debt":"1E8449","Liquid":"7D6608"}

    current_bucket = None
    for rank,f in enumerate(p['fund_plan'],1):
        bkt = f['bucket']

        # Bucket separator row
        if bkt != current_bucket:
            current_bucket = bkt
            ws.merge_cells(f"A{row}:L{row}")
            H(ws,row,1,f"  {bkt.upper()} FUNDS",bg=BUCKET_HDR.get(bkt,"1F4E79"),sz=9,al="left",colspan=1)
            ws.merge_cells(f"A{row}:L{row}")
            ws.row_dimensions[row].height = 16
            row += 1

        bg  = BUCKET_BG.get(bkt,"F8F9FA")
        is1 = (rank == 1)
        fbg = "D6EAF8" if (bkt=="Equity" and is1) else \
              "D5F5E3" if (bkt=="Debt"   and is1) else bg

        # Why this fund — short rationale
        why = build_rationale(f)

        vals = [
            bkt,
            f['category'],
            f['scheme_name'],
            f"{f['alloc_pct']}%",
            cr(f['monthly_sip']),
            fmt(f['return_1y'],"%",1),
            fmt(f['return_3y'],"%",1),
            fmt(f['return_5y'],"%",1),
            fmt(f['sharpe'],"",2),
            fmt(f['alpha'],"%",1),
            fmt(f['max_dd'],"%",1),
            why,
        ]

        for ci,v in enumerate(vals,1):
            al = "left" if ci in (2,3,12) else "center"
            c  = C(ws,row,ci,v,bg=fbg,sz=9,al=al,
                   bold=(ci==5),   # bold the SIP amount
                   wrap=(ci==12))
            if ci==5:  c.font = Font(color="1F4E79",bold=True,size=10,name="Arial")
            if ci==4:  c.font = Font(color="7D3C98",bold=True,size=9,name="Arial")
        ws.row_dimensions[row].height = 34 if any(len(str(v))>40 for v in vals) else 20
        row += 1

    # ── Totals row ────────────────────────────────────────────
    row += 1
    total_alloc = sum(f['alloc_pct'] for f in p['fund_plan'])
    total_plan_sip = sum(f['monthly_sip'] for f in p['fund_plan'])
    H(ws,row,1,"TOTAL",bg="0D1B2A",fg="FFFFFF",bold=True,sz=10,colspan=3)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
    C(ws,row,4,f"{total_alloc:.0f}%",bg="FFD700",fg="000000",bold=True,sz=10)
    C(ws,row,5,cr(total_plan_sip),bg="FFD700",fg="000000",bold=True,sz=11)
    ws.row_dimensions[row].height = 22

    # ── Year-by-year corpus projection ────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:L{row}")
    H(ws,row,1,"  YEAR-BY-YEAR CORPUS PROJECTION",bg="1F4E79",sz=11,al="left",colspan=1)
    ws.merge_cells(f"A{row}:L{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    ph = ["Year","Age","Invested","Equity Value","Debt Value","Total Corpus","vs Target","Milestone"]
    for ci,h in enumerate(ph,1):
        H(ws,row,ci,h,bg="2E4057",sz=9)
    row += 1

    milestones = {3:"3Y Review",5:"5Y Review",10:"10Y Review",g['time_years']:f"Goal: {g['name']}"}
    eq_r  = EXPECTED_EQUITY/12
    dt_r  = EXPECTED_DEBT/12
    ef    = p['eq_pct']/100
    df_   = p['debt_pct']/100
    sip   = g['monthly_sip']
    lump  = g['lumpsum']

    for yr in range(1,g['time_years']+1):
        m      = yr*12
        eq_v   = lump*ef*((1+EXPECTED_EQUITY)**yr) + (sip*ef*(((1+eq_r)**m-1)/eq_r*(1+eq_r)) if eq_r>0 else sip*ef*m)
        dt_v   = lump*df_*((1+EXPECTED_DEBT)**yr)  + (sip*df_*(((1+dt_r)**m-1)/dt_r*(1+dt_r)) if dt_r>0 else sip*df_*m)
        tot    = round(eq_v+dt_v)
        inv    = lump + sip*m
        ms     = milestones.get(yr,"")
        is_g   = yr==g['time_years']
        pct_tg = round(tot/g['target_amount']*100)
        bg     = "D1ECE1" if is_g else ("FFF9E6" if ms else "FAFAFA")

        for ci,v in enumerate([yr, CLIENT_PROFILE['age']+yr, cr(inv),
                                cr(round(eq_v)), cr(round(dt_v)),
                                cr(tot), f"{pct_tg}%", ms], 1):
            C(ws,row,ci,v,bg=bg,bold=is_g,sz=9)
        ws.row_dimensions[row].height = 16
        row += 1


def build_rationale(f):
    """Generate a short, clear reason why this fund is recommended."""
    parts = []
    r3 = f.get('return_3y')
    sh = f.get('sharpe')
    al = f.get('alpha')
    dd = f.get('max_dd')
    uc = f.get('upside_cap')
    dc = f.get('downside_cap')

    if r3 and not (isinstance(r3,float) and np.isnan(r3)):
        parts.append(f"{r3:.1f}% 3Y CAGR")
    if sh and not (isinstance(sh,float) and np.isnan(sh)) and float(sh) > 0.5:
        parts.append(f"Sharpe {sh:.2f} (excellent risk-adj return)")
    if al and not (isinstance(al,float) and np.isnan(al)) and float(al) > 1:
        parts.append(f"Alpha +{al:.1f}% vs benchmark")
    if dd and not (isinstance(dd,float) and np.isnan(dd)):
        parts.append(f"Max fall limited to {dd:.1f}%")
    if dc and not (isinstance(dc,float) and np.isnan(dc)) and float(dc) < 85:
        parts.append(f"Only {dc:.0f}% downside capture (defensive)")

    return " | ".join(parts[:3]) if parts else "Top composite score in category"


# ══════════════════════════════════════════════════════════════════
#  SHEET: MASTER FUND LIST (no duplicates, all goals)
# ══════════════════════════════════════════════════════════════════

def sheet_master(wb, portfolios):
    ws = wb.create_sheet("Master Fund List")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    ws["A1"].value = "  COMPLETE FUND LIST  —  All Goals Combined  |  No Fund Repeated"
    ws["A1"].fill  = PatternFill("solid",fgColor="0D1B2A")
    ws["A1"].font  = Font(color="FFD700",bold=True,size=13,name="Arial")
    ws["A1"].alignment = Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3
    heads = ["Goal","Fund Name","Category","Bucket","Alloc %",
             "Monthly SIP","3Y Return","5Y Return","Sharpe","Alpha %","Max DD","Rationale"]
    for ci,h in enumerate(heads,1):
        H(ws,row,ci,h,bg="1F4E79",sz=9)
    ws.row_dimensions[row].height = 20
    row += 1

    GOAL_COLORS = ["EBF5FB","EAFAF1","FEF9E7","F9EBEA","F5EEF8","FDFEFE"]
    for gi,p in enumerate(portfolios):
        gname = p['goal']['name']
        gc    = GOAL_COLORS[gi % len(GOAL_COLORS)]
        ws.merge_cells(f"A{row}:L{row}")
        H(ws,row,1,f"  {gname.upper()}",bg="2E4057",sz=9,al="left",colspan=1)
        ws.merge_cells(f"A{row}:L{row}")
        ws.row_dimensions[row].height = 16
        row += 1

        for f in p['fund_plan']:
            vals = [gname, f['scheme_name'], f['category'], f['bucket'],
                    f"{f['alloc_pct']}%", cr(f['monthly_sip']),
                    fmt(f['return_3y'],"%",1), fmt(f['return_5y'],"%",1),
                    fmt(f['sharpe'],"",2), fmt(f['alpha'],"%",1),
                    fmt(f['max_dd'],"%",1), build_rationale(f)]
            for ci,v in enumerate(vals,1):
                C(ws,row,ci,v,bg=gc,sz=9,
                  al="left" if ci in (1,2,3,12) else "center",
                  wrap=(ci==12))
            ws.row_dimensions[row].height = 20
            row += 1

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 14
    for c_ in ['G','H','I','J','K']:
        ws.column_dimensions[c_].width = 11
    ws.column_dimensions['L'].width = 34


# ══════════════════════════════════════════════════════════════════
#  SHEET: REBALANCING
# ══════════════════════════════════════════════════════════════════

def sheet_rebalancing(wb, portfolios):
    ws  = wb.create_sheet("Rebalancing Plan")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"].value = "  REBALANCING & REVIEW CALENDAR"
    ws["A1"].fill  = PatternFill("solid",fgColor="0D1B2A")
    ws["A1"].font  = Font(color="FFD700",bold=True,size=13,name="Arial")
    ws["A1"].alignment = Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3
    heads = ["Goal","Priority","Review Freq","Next Review","Trigger",
             "Action if Behind","De-risk From","Target Year","Notes"]
    for ci,h in enumerate(heads,1):
        H(ws,row,ci,h,bg="1F4E79",sz=9)
    ws.row_dimensions[row].height = 20
    row += 1

    RULES = {
        "Retirement":       ("Annual","Apr 2027","5% drift or >20% market fall","Increase SIP 10%","5 years before goal"),
        "ChildEducation":   ("Annual","Apr 2027","5% drift","Top up immediately","3 years before goal"),
        "TaxSaving":        ("Annual (March)","Mar 2027","March 31 deadline","Invest full 1.5L","Not applicable"),
        "HomePurchase":     ("Semi-annual","Oct 2026","7% drift","Lumpsum top-up","2 years before goal"),
        "WealthCreation":   ("Annual","Apr 2027","8% drift","Rebalance equity","—"),
        "EmergencyFund":    ("Quarterly","Aug 2026","Balance drops","Top up immediately","—"),
        "WealthPreservation":("Quarterly","Aug 2026","3% drift","Rebalance","—"),
    }

    for p in portfolios:
        g    = p['goal']
        r    = RULES.get(g['goal_type'],("Annual","Apr 2027","5% drift","Review","—"))
        hc,bgc,_ = PRIORITY_COLORS.get(g['priority'],("333333","F5F5F5","999999"))

        vals = [g['name'], g['priority'], r[0], r[1], r[2], r[3], r[4],
                str(date.today().year + g['time_years']),
                "Document changes | Consult advisor for large moves"]
        for ci,v in enumerate(vals,1):
            C(ws,row,ci,v,bg=bgc,sz=9,al="left" if ci in (1,5,6,9) else "center")
        ws.row_dimensions[row].height = 20
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:I{row}")
    H(ws,row,1,"  KEY RULES — SIP STEP-UP & SMART REBALANCING",bg="1F4E79",sz=11,al="left",colspan=1)
    ws.merge_cells(f"A{row}:I{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    rules = [
        ("SIP Step-Up",      "Increase SIP by 10-15% every April with salary increment. Most impactful action for long-term wealth."),
        ("Drift Rule",       "Rebalance if any asset class drifts more than 5% from target. Don't wait for annual review."),
        ("Market Crash Rule","If Nifty falls >20%: invest 3 months lumpsum in equity. Don't panic sell."),
        ("Market Rally Rule","If Nifty 1Y return >40%: book 10% profits. Move to debt or balanced advantage."),
        ("Life Events",      "Salary hike, new child, marriage, job change: Trigger full portfolio review immediately."),
        ("Fund Underperform","If any fund underperforms category by >3% for 2 consecutive years: Replace with next-ranked fund."),
        ("SIP Pause Rule",   "Never stop SIP for <6 months market correction. SIP works because of volatility, not despite it."),
    ]
    for label,desc in rules:
        H(ws,row,1,label,bg="2E4057",fg="A0B4C8",sz=9,al="left")
        ws.merge_cells(f"B{row}:I{row}")
        C(ws,row,2,f"  {desc}",bg="EEF5FF",sz=9,al="left",wrap=True)
        ws.merge_cells(f"B{row}:I{row}")
        ws.row_dimensions[row].height = 28
        row += 1

    for ci,w in enumerate([28,12,14,14,20,24,18,12,30],1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ══════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    print("="*65)
    print(f"  HNI GOALS MAPPER v2  |  {CLIENT_PROFILE['name']}  |  {CLIENT_PROFILE['risk_profile']}")
    print("="*65)

    if not os.path.exists(METRICS_CSV):
        print(f"❌ {METRICS_CSV} not found. Run main.py first!")
        return

    print(f"\n📊 Loading fund universe...")
    df = pd.read_csv(METRICS_CSV, dtype={'Scheme Code':str})
    if 'Category' not in df.columns: df['Category'] = 'Other'
    df['Category'] = df['Category'].fillna('Other')
    print(f"   {len(df):,} funds | {df['Category'].nunique()} categories")

    # Sort goals: MustHave first → best funds go to critical goals
    priority_order = {"MustHave":0,"ShouldHave":1,"Aspirational":2}
    sorted_goals   = sorted(GOALS, key=lambda g: priority_order.get(g['priority'],3))

    # Global tracker — ensures NO fund is repeated across goals
    used_codes = set()

    print(f"\n🎯 Building portfolios (no fund repeated across goals)...\n")
    portfolios = []
    # Restore original order after processing in priority order
    goal_order = {g['name']:i for i,g in enumerate(GOALS)}

    raw_portfolios = []
    for g in sorted_goals:
        print(f"   Processing: {g['name']} ({g['priority']})")
        p = build_portfolio(g, df, CLIENT_PROFILE['risk_profile'], used_codes)
        raw_portfolios.append(p)
        nf = len(p['fund_plan'])
        st = "✅ On Track" if p['on_track'] else "⚠ Top-Up"
        print(f"   → {nf} funds selected | Projected: {cr(p['projected'])} | {st}")
        print(f"   → Total unique funds used so far: {len(used_codes)}")
        print()

    # Restore original display order
    portfolios = sorted(raw_portfolios, key=lambda p: goal_order.get(p['goal']['name'],99))

    # Build Excel
    print(f"📊 Building Excel report...")
    wb = Workbook()
    sheet_dashboard(wb, CLIENT_PROFILE, portfolios)
    for i,p in enumerate(portfolios,1):
        sheet_goal(wb, p, i)
    sheet_master(wb, portfolios)
    sheet_rebalancing(wb, portfolios)

    wb.save(OUTPUT_FILE)

    print(f"\n{'='*65}")
    print(f"  ✅  DONE!")
    print(f"  📊  Open: {os.path.abspath(OUTPUT_FILE)}")
    print(f"  Total unique funds recommended: {len(used_codes)}")
    print(f"  No fund is repeated across any two goals.")
    print(f"{'='*65}\n")

    print(f"  Sheets:")
    print(f"   1. Dashboard         — All goals overview")
    for i,p in enumerate(portfolios,1):
        nf = len(p['fund_plan'])
        print(f"   {i+1}. {p['goal']['name']:<30} — {nf} funds with SIP amounts")
    print(f"   {len(portfolios)+2}. Master Fund List    — All funds, no duplicates")
    print(f"   {len(portfolios)+3}. Rebalancing Plan    — Annual review guide")


if __name__ == "__main__":
    main()
